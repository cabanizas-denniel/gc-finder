import pandas as pd
from datetime import datetime, timedelta
import random
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def generate_sample_lost_found_data():
    # Sample data for lost and found items
    items = ['Phone', 'Wallet', 'Keys', 'Laptop', 'Bag', 'ID Card', 'Watch', 'Umbrella', 'Water Bottle', 'Notebook']
    locations = ['Library', 'Cafeteria', 'Classroom A', 'Hallway', 'Gym', 'Auditorium', 'Main Entrance', 'Computer Lab']
    status_options = ['Found', 'Claimed', 'In Process', 'Returned to Owner']
    categories = ['Electronics', 'Personal Items', 'Documents', 'Accessories', 'School Supplies']
    
    num_records = 20
    today = datetime.now()
    
    data = {
        'Item_ID': [f'ITM{str(i+1).zfill(3)}' for i in range(num_records)],
        'Item_Name': [random.choice(items) for _ in range(num_records)],
        'Category': [random.choice(categories) for _ in range(num_records)],
        'Location_Found': [random.choice(locations) for _ in range(num_records)],
        'Date_Reported': [(today - timedelta(days=random.randint(0, 30))).strftime('%Y-%m-%d') for _ in range(num_records)],
        'Status': [random.choice(status_options) for _ in range(num_records)],
        'Description': ['Sample description for the found item' for _ in range(num_records)],
        'Contact_Person': [f'Person {i+1}' for i in range(num_records)]
    }
    
    return pd.DataFrame(data)

def generate_excel_report(filename='lost_and_found_report.xlsx'):
    # Get lost and found data
    df = generate_sample_lost_found_data()
    
    # Create Excel writer object
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Lost and Found Items', index=False, startrow=1)
        
        # Get the workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Lost and Found Items']
        
        # Add title
        worksheet['A1'] = 'GC Finder: Lost and Found Items Report'
        last_col = get_column_letter(len(df.columns))
        worksheet.merge_cells(f'A1:{last_col}1')
        title_cell = worksheet['A1']
        title_cell.font = Font(size=14, bold=True)
        
        # Format headers
        for col in range(len(df.columns)):
            cell = worksheet.cell(row=2, column=col+1)
            cell.font = Font(bold=True)
        
        # Add status summary
        status_summary = df['Status'].value_counts()
        summary_df = pd.DataFrame({'Status': status_summary.index, 'Count': status_summary.values})
        summary_df.to_excel(writer, sheet_name='Status Summary', index=False, startrow=1)
        
        # Format summary sheet
        summary_sheet = writer.sheets['Status Summary']
        summary_sheet['A1'] = 'Status Summary Report'
        summary_sheet.merge_cells('A1:B1')
        summary_sheet['A1'].font = Font(size=14, bold=True)
        
        # Adjust column widths in both sheets
        for sheet in [worksheet, summary_sheet]:
            for idx, col in enumerate(sheet.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet.column_dimensions[get_column_letter(idx)].width = max_length + 2

if __name__ == '__main__':
    generate_excel_report()
    print("Lost and Found report generated successfully!")
