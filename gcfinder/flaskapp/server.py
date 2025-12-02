import os
import sys
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import firebase_admin
from firebase_admin import credentials, firestore
import base64
from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json
import requests
from PIL import Image
import numpy as np
import traceback
import logging
from firebase_admin import auth
from functools import wraps
import re
import time

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize Flask app first
app = Flask(__name__)
# Limit incoming request size (e.g., ~5 MB) to mitigate large payload abuse
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024

# Initialize Firebase with robust error handling
def initialize_firebase():
    """Initialize Firebase with proper error handling"""
    try:
        if firebase_admin._apps:
            return firestore.client()
            
        cred_json_str = os.environ.get('FIREBASE_CREDENTIALS_JSON')
        if cred_json_str:
            try:
                cred_dict = json.loads(cred_json_str)
                cred = credentials.Certificate(cred_dict)
                logger.info("Firebase initialized with environment credentials")
            except json.JSONDecodeError as e:
                logger.error(f"Invalid Firebase credentials JSON: {e}")
                raise
        else:
            # Fallback to local file if environment variable is not set (for local development)
            credentials_file = "gcfinder-database-firebase-adminsdk-fbsvc-0447799241.json"
            if os.path.exists(credentials_file):
                cred = credentials.Certificate(credentials_file)
                logger.warning("Firebase initialized with local credentials file")
            else:
                raise FileNotFoundError("No Firebase credentials found. Set FIREBASE_CREDENTIALS_JSON environment variable or provide local credentials file.")

        firebase_admin.initialize_app(cred)
        return firestore.client()
    except Exception as e:
        logger.error(f"Failed to initialize Firebase: {e}")
        raise

# Initialize Firebase and database connection
try:
    db = initialize_firebase()
except Exception as e:
    logger.error(f"Critical error: Could not initialize Firebase: {e}")
    # For production, we might want to exit here
    db = None

# Hugging Face API configuration
HF_API_TOKEN = os.environ.get('HUGGING_FACE_API_TOKEN')
# Global switch to enable/disable remote HF inference
# Set DISABLE_HF_INFERENCE=1 in env to force local embeddings only
HF_AVAILABLE = os.environ.get('DISABLE_HF_INFERENCE', '').lower() not in ('1', 'true', 'yes')

def _to_fixed_embedding(vector, target_dim: int = 128) -> np.ndarray:
    """
    Convert any incoming embedding-like object to a fixed-size, normalized vector.
    - Flattens
    - Pads or truncates to target_dim
    - Replaces NaNs/Infs
    - L2-normalizes with epsilon guard
    """
    try:
        arr = np.array(vector, dtype=np.float32).reshape(-1)
    except Exception:
        arr = np.zeros(target_dim, dtype=np.float32)
    if arr.size == 0:
        arr = np.zeros(target_dim, dtype=np.float32)
    # Truncate or pad deterministically
    if arr.size > target_dim:
        arr = arr[:target_dim]
    elif arr.size < target_dim:
        # Repeat then cut to fit to preserve information order
        repeats = max(1, target_dim // max(1, arr.size))
        tiled = np.tile(arr, repeats)
        if tiled.size < target_dim:
            # Last resort pad zeros
            padded = np.zeros(target_dim, dtype=np.float32)
            padded[:tiled.size] = tiled
            arr = padded
        else:
            arr = tiled[:target_dim]
    # Clean NaNs/Infs
    arr = np.nan_to_num(arr, nan=0.0, posinf=0.0, neginf=0.0)
    # Normalize
    norm = float(np.linalg.norm(arr))
    if not np.isfinite(norm) or norm == 0.0:
        # Use a stable unit vector
        arr = np.zeros_like(arr)
        arr[0] = 1.0
    else:
        arr = arr / norm
    return arr

def validate_image_input(image_input):
    """Validate and standardize image input"""
    try:
        if isinstance(image_input, str) and image_input.startswith('data:image'):
            # Handle base64 data URL
            if ',' not in image_input:
                raise ValueError("Invalid base64 image format")
            image_data = image_input.split(',')[1]
            image_bytes = base64.b64decode(image_data)
            image = Image.open(BytesIO(image_bytes)).convert('RGB')
        elif isinstance(image_input, Image.Image):
            image = image_input.convert('RGB')
        else:
            # Handle file path or file-like object
            image = Image.open(image_input).convert('RGB')
        
        # Validate image size (prevent extremely large images)
        max_size = 4096  # 4K max
        if image.width > max_size or image.height > max_size:
            # Resize maintaining aspect ratio
            image.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
            logger.info(f"Resized large image to {image.size}")
            
        return image
    except Exception as e:
        logger.error(f"Error validating image input: {e}")
        raise ValueError(f"Invalid image format: {e}")

def get_local_image_embedding(image_input):
    """
    Fallback method to create simple image embeddings using basic image analysis
    when HuggingFace API is not available
    """
    try:
        image = validate_image_input(image_input)
        
        # Resize to standard size for consistency
        image = image.resize((64, 64), Image.Resampling.LANCZOS)
        
        # Convert to numpy array
        img_array = np.array(image, dtype=np.float32)
        
        # Extract basic features
        features = []
        
        # 1. Color histograms (RGB) - 48 features
        for channel in range(3):
            hist, _ = np.histogram(img_array[:,:,channel], bins=16, range=(0, 256))
            features.extend(hist.tolist())
        
        # 2. Average color values - 3 features
        avg_colors = np.mean(img_array, axis=(0, 1))
        features.extend(avg_colors.tolist())
        
        # 3. Color variance - 3 features
        var_colors = np.var(img_array, axis=(0, 1))
        features.extend(var_colors.tolist())
        
        # 4. Brightness and contrast - 2 features
        gray = np.mean(img_array, axis=2)
        brightness = np.mean(gray)
        contrast = np.std(gray)
        features.extend([brightness, contrast])
        
        # 5. Edge detection (simple gradient) - 1 feature
        grad_x = np.gradient(gray, axis=1)
        grad_y = np.gradient(gray, axis=0)
        edge_strength = np.mean(np.sqrt(grad_x**2 + grad_y**2))
        features.append(edge_strength)
        
        # 6. Texture features - 1 feature
        texture = np.std(gray)
        features.append(texture)
        
        # Total: 58 features - pad to 128
        # Convert to fixed-size normalized embedding
        embedding = _to_fixed_embedding(features, target_dim=128)
        logger.info(f"Created local embedding with {embedding.shape[0]} features")
        return embedding
        
    except Exception as e:
        logger.error(f"Error creating local embedding: {e}")
        # Return a normalized random embedding as last resort
        embedding = np.random.rand(128).astype(np.float32)
        return embedding / np.linalg.norm(embedding)

def try_classification_as_features(image_bytes):
    """Use ResNet for classification and extract feature-like scores"""
    global HF_AVAILABLE
    try:
        # Skip if remote inference disabled
        if not HF_AVAILABLE:
            return None
        headers = {"Content-Type": "application/json"}
        if HF_API_TOKEN:
            headers["Authorization"] = f"Bearer {HF_API_TOKEN}"
        
        # Convert to base64 for ResNet model
        image_b64 = base64.b64encode(image_bytes).decode('utf-8')
        
        # Use a reliable image classification model
        response = requests.post(
            "https://api-inference.huggingface.co/models/microsoft/resnet-50",
            headers=headers,
            json={"inputs": image_b64},
            timeout=45  # Increased timeout for reliability
        )
        
        logger.info(f"ResNet Classification Response Status: {response.status_code}")
        
        if response.status_code == 200:
            result = response.json()
            if isinstance(result, list) and len(result) > 0:
                # Extract scores to use as a simple embedding
                scores = []
                for item in result:
                    if isinstance(item, dict) and 'score' in item:
                        scores.append(float(item['score']))
                
                if len(scores) >= 5:  # Need enough features
                    # Pad to make a reasonable embedding size
                    while len(scores) < 128:
                        scores.extend(scores[:min(len(scores), 128-len(scores))])
                    
                    embedding = np.array(scores[:128], dtype=np.float32)
                    
                    # Normalize to prevent numerical issues
                    norm = np.linalg.norm(embedding)
                    if norm > 0:
                        embedding = embedding / norm
                    
                    logger.info(f"Successfully got ResNet-based embedding with shape: {embedding.shape}")
                    return embedding
        else:
            logger.warning(f"ResNet classification failed: {response.status_code} - {response.text[:200]}")
            if response.status_code == 410:
                # Permanently disable remote inference for this process to avoid repeated 410s
                HF_AVAILABLE = False
            
    except requests.exceptions.Timeout:
        logger.error("ResNet classification timed out")
    except requests.exceptions.RequestException as e:
        logger.error(f"ResNet classification network error: {e}")
    except Exception as e:
        logger.error(f"ResNet classification error: {e}")
        
    return None

def try_sentence_transformers_clip(image_bytes):
    """Try using sentence-transformers CLIP model"""
    global HF_AVAILABLE
    try:
        # Skip if remote inference disabled
        if not HF_AVAILABLE:
            return None
        headers = {}
        if HF_API_TOKEN:
            headers["Authorization"] = f"Bearer {HF_API_TOKEN}"
        
        response = requests.post(
            "https://api-inference.huggingface.co/models/sentence-transformers/clip-ViT-B-32",
            headers=headers,
            data=image_bytes,
            timeout=45
        )
        
        logger.info(f"Sentence-transformers CLIP Response Status: {response.status_code}")
        
        if response.status_code == 200:
            result = response.json()
            if isinstance(result, list) and len(result) > 0:
                embedding = _to_fixed_embedding(result, target_dim=128)
                logger.info(f"Successfully got sentence-transformers embedding with shape: {embedding.shape}")
                return embedding
        else:
            logger.warning(f"Sentence-transformers CLIP failed: {response.status_code}")
            if response.status_code == 410:
                HF_AVAILABLE = False
            
    except Exception as e:
        logger.error(f"Sentence-transformers CLIP error: {e}")
        
    return None

def try_openai_clip_feature_extraction(image_bytes):
    """Try using OpenAI CLIP model for feature extraction"""
    global HF_AVAILABLE
    try:
        # Skip if remote inference disabled
        if not HF_AVAILABLE:
            return None
        headers = {}
        if HF_API_TOKEN:
            headers["Authorization"] = f"Bearer {HF_API_TOKEN}"
        
        response = requests.post(
            "https://api-inference.huggingface.co/models/openai/clip-vit-base-patch32",
            headers=headers,
            data=image_bytes,
            timeout=45
        )
        
        logger.info(f"OpenAI CLIP Feature Extraction Response Status: {response.status_code}")
        
        if response.status_code == 200:
            result = response.json()
            if isinstance(result, list) and len(result) > 0:
                if isinstance(result[0], list):
                    raw = result[0]
                else:
                    raw = result
                embedding = _to_fixed_embedding(raw, target_dim=128)
                logger.info(f"Successfully got OpenAI CLIP embedding with shape: {embedding.shape}")
                return embedding
        else:
            logger.warning(f"OpenAI CLIP feature extraction failed: {response.status_code}")
            if response.status_code == 410:
                HF_AVAILABLE = False
            
    except Exception as e:
        logger.error(f"OpenAI CLIP feature extraction error: {e}")
        
    return None

def get_image_embedding_remote(image_input):
    """
    Get image embedding using HuggingFace's models with robust fallbacks
    """
    try:
        # If remote is disabled, use local immediately
        if not HF_AVAILABLE:
            return _to_fixed_embedding(get_local_image_embedding(image_input), target_dim=128)
        # Validate and convert image to bytes
        image = validate_image_input(image_input)
        buffer = BytesIO()
        image.save(buffer, format='JPEG', quality=85, optimize=True)
        image_bytes = buffer.getvalue()
        
        # Try methods in order of preference
        methods = [
            ("ResNet Classification", try_classification_as_features),
            ("Sentence Transformers CLIP", try_sentence_transformers_clip),
            ("OpenAI CLIP", try_openai_clip_feature_extraction)
        ]
        
        for method_name, method_func in methods:
            try:
                logger.info(f"Trying {method_name}...")
                result = method_func(image_bytes)
                if result is not None:
                    fixed = _to_fixed_embedding(result, target_dim=128)
                    logger.info(f"Successfully got embedding using {method_name} -> fixed dim {fixed.shape[0]}")
                    return fixed
            except Exception as e:
                logger.warning(f"{method_name} failed: {e}")
                continue
        
        logger.warning("All HuggingFace API methods failed, falling back to local embedding...")
        return _to_fixed_embedding(get_local_image_embedding(image_input), target_dim=128)
            
    except Exception as e:
        logger.error(f"Error in get_image_embedding_remote: {e}")
        traceback.print_exc()
        
        # Last resort: try local embedding
        try:
            return _to_fixed_embedding(get_local_image_embedding(image_input), target_dim=128)
        except Exception as local_e:
            logger.error(f"Local embedding also failed: {local_e}")
            # Return normalized random embedding as absolute last resort
            embedding = np.random.rand(128).astype(np.float32)
            return embedding / np.linalg.norm(embedding)

def compute_similarity(query_embedding, database_embeddings):
    """
    Compute cosine similarity between query and database embeddings with error handling
    """
    try:
        if len(database_embeddings) == 0:
            return np.array([])
            
        # Ensure all embeddings are same dimension
        query_flat = _to_fixed_embedding(query_embedding, target_dim=128)
        db_list = [ _to_fixed_embedding(emb, target_dim=128) for emb in database_embeddings ]
        db_flat = np.stack(db_list, axis=0)  # shape: (N, 128)
        
        # Normalize embeddings
        eps = 1e-8
        qnorm = float(np.linalg.norm(query_flat))
        qnorm = qnorm if (np.isfinite(qnorm) and qnorm > 0.0) else eps
        query_norm = query_flat / qnorm
        db_norms_den = np.linalg.norm(db_flat, axis=1, keepdims=True)
        db_norms_den = np.where(np.isfinite(db_norms_den) & (db_norms_den > 0.0), db_norms_den, eps)
        db_norms = db_flat / db_norms_den
        
        # Compute cosine similarity
        similarities = np.dot(db_norms, query_norm)
        
        # Ensure similarities are in valid range [-1, 1]
        similarities = np.clip(similarities, -1.0, 1.0)
        
        return similarities
    except Exception as e:
        logger.error(f"Error computing similarity: {e}")
        return np.array([])

# Configure CORS with production-ready settings
CORS(app, resources={
    r"/*": {
        "origins": [
            "https://gc-finder.vercel.app", 
            "https://gcfinder.pages.dev", 
            "http://localhost:3000",
            "https://localhost:3000",
            "http://127.0.0.1:3000"
        ],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "supports_credentials": False,
        "expose_headers": ["Content-Type", "Authorization"]
    }
})

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Allow CORS preflight to pass
        if request.method == 'OPTIONS':
            return ('', 200)
        id_token = request.headers.get('Authorization')
        if not id_token or not id_token.startswith('Bearer '):
            return jsonify({'error': 'Authorization token is missing or invalid'}), 401

        id_token = id_token.split('Bearer ')[1]

        try:
            # Verify the token against the Firebase Auth API.
            decoded_token = auth.verify_id_token(id_token)
            uid = decoded_token['uid']

            # Check if the user is an admin by looking them up in the 'admin' collection.
            admin_ref = db.collection('admin').document(uid)
            admin_doc = admin_ref.get()
            if not admin_doc.exists:
                return jsonify({'error': 'User is not an admin'}), 403
            
            # Add user to request context
            request.user = decoded_token

        except auth.InvalidIdTokenError:
            return jsonify({'error': 'Invalid ID token'}), 401
        except Exception as e:
            logger.error(f"Token verification error: {e}")
            return jsonify({'error': 'An internal error occurred during authentication'}), 500

        return f(*args, **kwargs)
    return decorated_function

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Allow CORS preflight to pass
        if request.method == 'OPTIONS':
            return ('', 200)
        id_token = request.headers.get('Authorization')
        if not id_token or not id_token.startswith('Bearer '):
            return jsonify({'error': 'Authorization token is missing or invalid'}), 401

        id_token = id_token.split('Bearer ')[1]

        try:
            decoded_token = auth.verify_id_token(id_token)
            request.user = decoded_token
        except auth.InvalidIdTokenError:
            return jsonify({'error': 'Invalid ID token'}), 401
        except Exception as e:
            logger.error(f"Token verification error: {e}")
            return jsonify({'error': 'An internal error occurred during authentication'}), 500

        return f(*args, **kwargs)
    return decorated_function

# Health check endpoint
@app.route("/")
def home():
    """Health check endpoint"""
    db_status = "connected" if db is not None else "disconnected"
    hf_token_status = "set" if HF_API_TOKEN else "not_set"
    
    return jsonify({
        "status": "ok", 
        "message": "GCFinder API is running!",
        "database": db_status,
        "hf_token": hf_token_status,
        "version": "1.0.0"
    }), 200

@app.route('/test-api', methods=['GET'])
def test_api():
    """Test endpoint to check HuggingFace API connectivity and model availability"""
    try:
        if not db:
            return jsonify({
                'overall_status': 'error',
                'error': 'Database not connected'
            }), 500
            
        # Test with a simple test image
        test_image = Image.new('RGB', (32, 32), color='red')
        buffer = BytesIO()
        test_image.save(buffer, format='JPEG')
        image_bytes = buffer.getvalue()
        
        logger.info("Testing HuggingFace API connectivity...")
        
        # Try each method
        results = {}
        
        # Test ResNet (most reliable)
        result = try_classification_as_features(image_bytes)
        results['resnet_classification'] = {
            'success': result is not None,
            'shape': result.shape if result is not None else None
        }
        
        # Test sentence-transformers CLIP
        result = try_sentence_transformers_clip(image_bytes)
        results['sentence_transformers_clip'] = {
            'success': result is not None,
            'shape': result.shape if result is not None else None
        }
        
        # Test OpenAI CLIP
        result = try_openai_clip_feature_extraction(image_bytes)
        results['openai_clip'] = {
            'success': result is not None,
            'shape': result.shape if result is not None else None
        }
        
        # Test local fallback
        result = get_local_image_embedding(test_image)
        results['local_fallback'] = {
            'success': result is not None,
            'shape': result.shape if result is not None else None
        }
        
        # Check overall status
        any_success = any(r['success'] for r in results.values())
        
        return jsonify({
            'overall_status': 'success' if any_success else 'failed',
            'hf_token_set': HF_API_TOKEN is not None,
            'database_connected': db is not None,
            'detailed_results': results,
            'message': 'At least one method works' if any_success else 'All methods failed - check logs'
        })
        
    except Exception as e:
        logger.error(f"API test error: {e}")
        traceback.print_exc()
        return jsonify({
            'overall_status': 'error',
            'error': str(e)
        }), 500

@app.route('/api/batch-create-students', methods=['POST'])
@admin_required
def batch_create_students():
    """
    Creates student users in Firebase Auth and corresponding documents in Firestore.
    """
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    students_data = request.get_json().get('students')
    if not students_data or not isinstance(students_data, list):
        return jsonify({"error": "Missing or invalid 'students' array in request body"}), 400
    
    success_count = 0
    failure_count = 0
    results = []

    for student in students_data:
        email = student.get('email')
        password = student.get('password')
        full_name = student.get('full_name')
        student_id = student.get('student_id')
        year_level = student.get('year_level')
        status = student.get('status', 'active')

        if not all([email, password, full_name, student_id, year_level]):
            failure_count += 1
            results.append({'email': email, 'status': 'failed', 'reason': 'Missing required fields.'})
            continue

        try:
            # Create user in Firebase Authentication with student_id as UID
            user_record = auth.create_user(
                uid=student_id,  # Use student_id as UID for consistency
                email=email,
                password=password,
                display_name=full_name,
                email_verified=True
            )
            
            # Create user document in Firestore with the same UID (student_id)
            student_doc_ref = db.collection('students').document(student_id)
            student_doc_ref.set({
                'full_name': full_name,
                'student_id': student_id,
                'email': email,
                'year_level': year_level,
                'status': status,
                'createdAt': firestore.SERVER_TIMESTAMP
            })
            
            success_count += 1
            results.append({'email': email, 'status': 'success', 'uid': user_record.uid})

        except Exception as e:
            failure_count += 1
            error_reason = str(e)
            results.append({'email': email, 'status': 'failed', 'reason': error_reason})

    return jsonify({
        "message": "Batch creation process completed.",
        "success_count": success_count,
        "failure_count": failure_count,
        "results": results
    }), 200

@app.route('/api/batch-create-users', methods=['POST'])
@admin_required
def batch_create_users():
    """
    Creates users (students or staff) in Firebase Auth and Firestore.
    Accepts an array 'users' where each entry can include:
      student_id (used as UID), full_name, email, password, status, role, year_level (optional for staff)
    """
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    users_data = request.get_json().get('users')
    if not users_data or not isinstance(users_data, list):
        return jsonify({"error": "Missing or invalid 'users' array in request body"}), 400

    valid_roles = {'student', 'official'}  # 'official' for faculty/staff

    success_count = 0
    failure_count = 0
    results = []

    def generate_password(email_addr, user_role, full_name_str):
        try:
            username = (email_addr or '').split('@')[0]
            if (user_role or '').lower() == 'student':
                digits = ''.join([c for c in username if c.isdigit()])
                hash_part = (digits[-5:] if digits else username[-5:]).rjust(5, '0')
            else:
                # Official: try to use last name first
                hash_part = ''
                if full_name_str:
                    parts = full_name_str.strip().split()
                    if parts:
                        last_name = parts[-1]
                        # Keep only letters for the password part
                        hash_part = ''.join([c for c in last_name if c.isalpha()])
                
                # Fallback to email base if last name didn't work
                if not hash_part:
                    base = username.split('.')[0] if username else ''
                    hash_part = ''.join([c for c in base if c.isalpha()]) or 'user'

            year = datetime.now().year
            return f"GC{hash_part}{year}"
        except Exception:
            # Fallback if parsing fails
            return f"GCuser{datetime.now().year}"

    for user in users_data:
        email = user.get('email')
        # password is optional/ignored; server will generate based on rules
        full_name = user.get('full_name')
        student_id = user.get('student_id')
        year_level = user.get('year_level')
        status = user.get('status', 'active')
        role = (user.get('role') or 'student').lower()

        if role not in valid_roles:
            failure_count += 1
            results.append({'email': email, 'status': 'failed', 'reason': f"Invalid role '{role}'."})
            continue

        # For students, require year_level
        if role == 'student' and (year_level is None or str(year_level).strip() == ''):
            failure_count += 1
            results.append({'email': email, 'status': 'failed', 'reason': 'Missing year_level for student.'})
            continue

        # Default email if missing (student_id based)
        if not email and student_id:
            email = f"{student_id}@gordoncollege.edu.ph"

        if not all([email, full_name, student_id]):
            failure_count += 1
            results.append({'email': email, 'status': 'failed', 'reason': 'Missing required fields.'})
            continue

        try:
            gen_password = generate_password(email, role, full_name)
            # Create user in Firebase Authentication with student_id/employee_id as UID
            user_record = auth.create_user(
                uid=student_id,
                email=email,
                password=gen_password,
                display_name=full_name,
                email_verified=True
            )

            # Determine which collection to use based on role
            # 'official' goes to 'officials' collection, 'student' goes to 'students'
            is_official = role == 'official'
            collection_name = 'officials' if is_official else 'students'
            
            # Create user document in Firestore with the same UID
            user_doc_ref = db.collection(collection_name).document(student_id)
            doc_payload = {
                'full_name': full_name,
                'email': email,
                'status': status,
                'role': role,
                'createdAt': firestore.SERVER_TIMESTAMP
            }
            
            # Add appropriate ID field based on role
            if is_official:
                doc_payload['employee_id'] = student_id
            else:
                doc_payload['student_id'] = student_id
                if year_level is not None:
                    doc_payload['year_level'] = year_level

            user_doc_ref.set(doc_payload)

            success_count += 1
            results.append({'email': email, 'status': 'success', 'uid': user_record.uid, 'collection': collection_name})

        except Exception as e:
            failure_count += 1
            error_reason = str(e)
            results.append({'email': email, 'status': 'failed', 'reason': error_reason})

    return jsonify({
        "message": "Batch creation process completed.",
        "success_count": success_count,
        "failure_count": failure_count,
        "results": results
    }), 200

@app.route('/api/users/<uid>/role', methods=['PUT'])
@admin_required
def update_user_role(uid):
    """Update a user's role (student or staff) - Admin only"""
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    role = (request.get_json() or {}).get('role')
    if not role:
        return jsonify({"error": "Missing 'role' in request body"}), 400

    role = str(role).lower()
    if role not in {'student', 'staff'}:
        return jsonify({"error": "Invalid role. Must be 'student' or 'staff'"}), 400

    try:
        user_ref = db.collection('students').document(uid)
        user_doc = user_ref.get()
        if not user_doc.exists:
            return jsonify({"error": "User not found"}), 404

        user_ref.update({
            'role': role,
            'updatedAt': firestore.SERVER_TIMESTAMP
        })

        return jsonify({
            'message': f"User {uid} role updated to {role}",
            'role': role
        }), 200
    except Exception as e:
        logger.error(f"Error updating user role: {e}")
        return jsonify({"error": "Failed to update user role"}), 500

@app.route('/api/delete-user', methods=['DELETE'])
@admin_required
def delete_user():
    """
    Deletes a user from Firebase Auth and their corresponding document from Firestore.
    Checks both students and officials collections.
    """
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    uid = request.get_json().get('uid')
    if not uid:
        return jsonify({"error": "Missing 'uid' in request body"}), 400

    try:
        # Delete user from Firebase Authentication
        auth.delete_user(uid)
        
        # Try to delete from students collection
        student_doc_ref = db.collection('students').document(uid)
        if student_doc_ref.get().exists:
            student_doc_ref.delete()
        
        # Also try to delete from officials collection
        official_doc_ref = db.collection('officials').document(uid)
        if official_doc_ref.get().exists:
            official_doc_ref.delete()
        
        return jsonify({"message": f"Successfully deleted user {uid}"}), 200

    except auth.UserNotFoundError:
        return jsonify({"error": f"User with UID {uid} not found in Firebase Authentication."}), 404
    except Exception as e:
        logger.error(f"Error deleting user {uid}: {e}")
        return jsonify({"error": "An internal error occurred while deleting the user."}), 500

@app.route('/api/change-password', methods=['POST'])
@login_required
def change_password():
    """
    Changes the user's password in Firebase Authentication.
    Note: This endpoint does not verify the user's current password.
    For security, re-authentication should be handled on the client-side
    before calling this endpoint.
    """
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    new_password = request.get_json().get('newPassword')
    if not new_password or len(new_password) < 6:
        return jsonify({"error": "New password must be at least 6 characters long"}), 400

    uid = request.user['uid']

    try:
        auth.update_user(uid, password=new_password)
        return jsonify({"message": "Password updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error changing password for user {uid}: {e}")
        return jsonify({"error": "An internal error occurred while changing the password."}), 500

@app.route('/search', methods=['POST'])
def search():
    """Enhanced search endpoint with comprehensive error handling"""
    
    # Validate database connection
    if not db:
        return jsonify({'error': 'Database not available'}), 503
        
    # Validate file upload
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    try:
        logger.info(f"Processing search request with file: {file.filename}")
        
        # Log API token status
        if not HF_API_TOKEN:
            logger.warning("HUGGING_FACE_API_TOKEN not set. Using public API with rate limits.")
        else:
            logger.info("Using authenticated HuggingFace API")
        
        # Process query image with validation
        logger.info("Processing query image...")
        try:
            query_image = Image.open(file).convert('RGB')
            logger.info(f"Query image size: {query_image.size}")
        except Exception as e:
            logger.error(f"Invalid image file: {e}")
            return jsonify({'error': 'Invalid image file format'}), 400
        
        # Get image embedding
        logger.info("Getting image embedding...")
        query_embedding = get_image_embedding_remote(query_image)
        
        if query_embedding is None:
            logger.error("Failed to get query embedding")
            return jsonify({'error': 'Failed to process query image. Please try again later.'}), 503

        logger.info(f"Successfully got query embedding with shape: {query_embedding.shape}")

        # Fetch and process database items with error handling
        logger.info("Fetching items from database...")
        try:
            items_ref = db.collection('items')
            items = items_ref.where('status', '!=', 'archived').stream()
        except Exception as e:
            logger.error(f"Database query failed: {e}")
            return jsonify({'error': 'Database query failed'}), 500
        
        database_items = []
        database_embeddings = []
        processed_items = 0
        skipped_items = 0
        
        for item in items:
            try:
                item_data = item.to_dict()
                
                # Apply filtering logic
                if item_data.get('adminApproval') != True:
                    skipped_items += 1
                    continue
                    
                # Only process items that have image data
                if 'imageData' in item_data and item_data['imageData'] and len(item_data['imageData']) > 0:
                    try:
                        image_data = item_data['imageData'][0]['dataUrl']
                        if not image_data:
                            skipped_items += 1
                            continue
                            
                        embedding = get_image_embedding_remote(image_data)
                        
                        if embedding is not None:
                            database_embeddings.append(embedding.flatten())
                            database_items.append({
                                'id': item.id,
                                'name': item_data.get('name', 'Unnamed Item'),
                                'category': item_data.get('category', 'Uncategorized'),
                                'location': item_data.get('location', 'Unknown location'),
                                'date': item_data.get('date', ''),
                                'description': item_data.get('description', ''),
                                'status': item_data.get('status', 'Unclaimed'),
                                'image': image_data,
                                'submitter': item_data.get('submitter', None),
                                'adminApproval': item_data.get('adminApproval', False)
                            })
                            processed_items += 1
                        else:
                            logger.warning(f"Failed to get embedding for item {item.id}")
                            skipped_items += 1
                    except Exception as e:
                        logger.warning(f"Error processing item {item.id}: {e}")
                        skipped_items += 1
                        continue
                else:
                    skipped_items += 1
                    
            except Exception as e:
                logger.warning(f"Error processing item document: {e}")
                skipped_items += 1
                continue

        logger.info(f"Processed {processed_items} items, skipped {skipped_items} items")

        if not database_items:
            return jsonify({'error': 'No approved items with images found in database'}), 404

        # Compute and sort similarities
        logger.info("Computing similarities...")
        try:
            similarities = compute_similarity(query_embedding.flatten(), database_embeddings)
            
            if len(similarities) == 0:
                return jsonify({'error': 'Failed to compute similarities'}), 500
                
            # Sort strictly by similarity score to avoid comparing dicts on ties
            order = np.argsort(similarities)[::-1]
            results = [(float(similarities[i]), database_items[i]) for i in order]
            
            logger.info(f"Returning {len(results)} results")
            return jsonify({
                'results': [
                    {
                        'item': item,
                        'similarity': float(score)
                    }
                    for score, item in results
                ],
                'processed_items': processed_items,
                'total_results': len(results)
            })
            
        except Exception as e:
            logger.error(f"Error computing similarities: {e}")
            return jsonify({'error': 'Failed to compute similarities'}), 500
        
    except Exception as e:
        logger.error(f"Search error: {e}")
        traceback.print_exc()
        return jsonify({'error': f'Search failed: {str(e)}'}), 500

def fetch_items_for_export(start_date_dt=None, end_date_dt=None):
    """
    Fetches item data from Firestore for Excel export, filtered by date in Python.
    Handles various date formats stored in Firestore.
    """
    if not db:
        logger.error("Database not available for export")
        return pd.DataFrame()
        
    try:
        items_ref = db.collection('items')
        docs = items_ref.stream()
        data = []

        # Prepare date objects for filtering, using only the date part
        filter_start_date = start_date_dt.date() if start_date_dt else None
        filter_end_date = end_date_dt.date() if end_date_dt else None

        for doc in docs:
            try:
                item_data = doc.to_dict()
                
                doc_date_obj = None  # This will be a datetime.date object if parsing succeeds
                doc_date_raw = item_data.get('date')
                date_reported_str = 'N/A'  # For display in Excel

                if doc_date_raw:
                    actual_datetime_obj = None
                    if isinstance(doc_date_raw, datetime):  # Already a datetime (e.g., from Firestore Timestamp)
                        actual_datetime_obj = doc_date_raw
                    elif isinstance(doc_date_raw, str):
                        # Try parsing various common date and datetime formats
                        for fmt in (
                            '%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%m-%d-%Y', '%Y/%m/%d', # Common date formats
                            '%m/%d/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', # Common datetime formats
                            '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%dT%H:%M:%SZ', # ISO formats
                            '%a, %d %b %Y %H:%M:%S %Z' # Example: 'Mon, 12 Jul 2021 14:30:00 GMT' (RFC 822/1123) - less likely from UI
                        ):
                            try:
                                actual_datetime_obj = datetime.strptime(doc_date_raw, fmt)
                                break  # Stop on first successful parse
                            except ValueError:
                                continue
                    
                    if actual_datetime_obj:
                        doc_date_obj = actual_datetime_obj.date()  # Use date part for filtering
                        date_reported_str = actual_datetime_obj.strftime('%Y-%m-%d %H:%M:%S')  # Use full datetime for display
                    elif isinstance(doc_date_raw, str): # If parsing failed, keep original string for display
                        date_reported_str = doc_date_raw
                    else: # If some other type (e.g. number if mistake was made), convert to string for display
                        date_reported_str = str(doc_date_raw)

                # Python-based date filtering
                if filter_start_date and (not doc_date_obj or doc_date_obj < filter_start_date):
                    continue  # Skip item if its date is before start_date
                if filter_end_date and (not doc_date_obj or doc_date_obj > filter_end_date):
                    continue  # Skip item if its date is after end_date

                # Handle potentially nested submitter information
                submitter_map = item_data.get('submitter') # Get the 'submitter' object/map
                
                submitter_user_id = None
                full_name = 'N/A'
                student_id = 'N/A'

                if isinstance(submitter_map, dict):
                    # Try to get userId and denormalized details directly from the submitter_map
                    submitter_user_id = submitter_map.get('userId') # Assumes userId is within the submitter map
                    full_name = submitter_map.get('displayName', submitter_map.get('full_name', 'N/A'))
                    student_id = submitter_map.get('studentId', submitter_map.get('student_id', 'N/A'))
                else:
                    submitter_user_id = item_data.get('userId', item_data.get('submitter')) 

                if submitter_user_id and isinstance(submitter_user_id, str) and submitter_user_id != 'N/A':
                    if full_name == 'N/A' or student_id == 'N/A': 
                        try:
                            user_doc_ref = db.collection('users').document(submitter_user_id)
                            user_doc = user_doc_ref.get()
                            if user_doc.exists:
                                user_details = user_doc.to_dict()
                                # Override with fetched data if it was N/A from submitter_map
                                if full_name == 'N/A':
                                    full_name = user_details.get('displayName', user_details.get('name', 'N/A'))
                                if student_id == 'N/A':
                                    student_id = user_details.get('studentId', user_details.get('studentID', 'N/A'))
                            else:
                                logger.warning(f"User document not found for ID: {submitter_user_id} (item {doc.id})")
                        except Exception as e:
                            logger.warning(f"Error fetching user details for {submitter_user_id} (item {doc.id}): {e}")
                elif submitter_user_id: # Catches cases where submitter_user_id might be non-string after initial direct get if not map
                     logger.warning(f"Invalid or non-string submitter_user_id found: {submitter_user_id} for item {doc.id}")

                data.append({
                    'Item_ID': doc.id,
                    'Item_Name': item_data.get('name', 'N/A'),
                    'Category': item_data.get('category', 'N/A'),
                    'Location_Found': item_data.get('location', 'N/A'),
                    'Submitted_At': date_reported_str, 
                    'Status': item_data.get('status', 'N/A'),
                    'Description': item_data.get('description', 'N/A'),
                    'Student_ID': student_id, 
                    'Full_Name': full_name,    
                    'Admin_Approved': item_data.get('adminApproval', False)
                })
            except Exception as e:
                logger.warning(f"Error processing item {doc.id} for export: {e}")
                continue
        
        if not data:
            return pd.DataFrame(columns=['Item_ID', 'Item_Name', 'Category', 'Location_Found', 'Submitted_At', 'Status', 'Description', 'Student_ID', 'Full_Name', 'Admin_Approved'])
        return pd.DataFrame(data)
    except Exception as e:
        logger.error(f"Error fetching items for export: {e}")
        return pd.DataFrame()

def fetch_users_for_export(start_date_dt=None, end_date_dt=None):
    """
    Fetches user data from Firestore for Excel export.
    Date filtering is currently disabled as 'students' documents lack a date field.
    """
    if not db:
        logger.error("Database not available for export")
        return pd.DataFrame()
        
    try:
        users_ref = db.collection('students') 
        query = users_ref
            
        docs = query.stream()
        data = []
        for doc in docs:
            try:
                user = doc.to_dict()
                user_data_for_df = {
                    'User_ID': doc.id,
                    'Name': user.get('full_name', user.get('name', 'N/A')), 
                    'Email': user.get('email', 'N/A'),
                    'Student_ID': user.get('student_id', 'N/A'),
                    'Year_Level': user.get('year_level', 'N/A'),
                    'Status': user.get('status', 'active')
                }
                data.append(user_data_for_df)
            except Exception as e:
                logger.warning(f"Error processing user {doc.id} for export: {e}")
                continue
            
        # Define column order for the DataFrame
        df_columns = ['User_ID', 'Name', 'Email', 'Student_ID', 'Year_Level', 'Status']
        
        if not data:
            return pd.DataFrame(columns=df_columns)
        return pd.DataFrame(data, columns=df_columns)
    except Exception as e:
        logger.error(f"Error fetching users for export: {e}")
        return pd.DataFrame()

def generate_excel_report_to_buffer(data_type, start_date_str=None, end_date_str=None):
    """Generate Excel report with error handling"""
    try:
        start_date_dt = None
        end_date_dt = None
        if start_date_str:
            try:
                start_date_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
            except ValueError:
                # Allow 'None' or empty string to pass through as no date filter
                if start_date_str.lower() != 'none' and start_date_str != '':
                    raise ValueError(f"Invalid start_date format: {start_date_str}. Expected YYYY-MM-DD.")
        if end_date_str:
            try:
                end_date_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
            except ValueError:
                if end_date_str.lower() != 'none' and end_date_str != '':
                    raise ValueError(f"Invalid end_date format: {end_date_str}. Expected YYYY-MM-DD.")

        if data_type == 'items':
            df = fetch_items_for_export(start_date_dt, end_date_dt)
            main_sheet_name = 'Items Report'
            title = 'GC Finder: Items Report'
            summary_field = 'Status' 
            summary_title = 'Item Status Summary'
        elif data_type == 'users':
            df = fetch_users_for_export(start_date_dt, end_date_dt) # Dates are now effectively ignored by fetch_users_for_export
            main_sheet_name = 'Users Report'
            title = 'GC Finder: Users Report'
            summary_field = 'Status' 
            summary_title = 'User Status Summary'
        else:
            return None

        if df.empty:
            output_buffer = BytesIO()
            empty_df_message = pd.DataFrame([{'message': f'No {data_type} found for the selected criteria.'}])
            with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                empty_df_message.to_excel(writer, sheet_name='No Data', index=False)
            output_buffer.seek(0)
            return output_buffer
            
        output_buffer = BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=main_sheet_name, index=False, startrow=1)
            
            workbook = writer.book
            worksheet = writer.sheets[main_sheet_name]
            
            worksheet['A1'] = title
            last_col_letter = get_column_letter(len(df.columns) if not df.empty else 1)
            worksheet.merge_cells(f'A1:{last_col_letter}1')
            title_cell = worksheet['A1']
            title_cell.font = Font(size=14, bold=True, name='Calibri')
            title_cell.alignment = Alignment(horizontal='center')
            
            for col_idx, column_title in enumerate(df.columns):
                cell = worksheet.cell(row=2, column=col_idx + 1)
                cell.value = column_title
                cell.font = Font(bold=True, name='Calibri')
                cell.alignment = Alignment(horizontal='center')

            status_col_name = summary_field
            if not df.empty and status_col_name in df.columns and not df[status_col_name].dropna().empty:
                status_summary_df = df[status_col_name].value_counts().reset_index()
                status_summary_df.columns = [status_col_name, 'Count']
                
                summary_sheet_name = summary_title
                status_summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False, startrow=1)
                
                summary_ws = writer.sheets[summary_sheet_name]
                summary_ws['A1'] = f"{main_sheet_name} - {summary_title}"
                summary_ws.merge_cells('A1:B1')
                summary_ws['A1'].font = Font(size=14, bold=True, name='Calibri')
                summary_ws['A1'].alignment = Alignment(horizontal='center')

                summary_ws.cell(row=2, column=1).value = status_col_name
                summary_ws.cell(row=2, column=1).font = Font(bold=True, name='Calibri')
                summary_ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
                summary_ws.cell(row=2, column=2).value = 'Count'
                summary_ws.cell(row=2, column=2).font = Font(bold=True, name='Calibri')
                summary_ws.cell(row=2, column=2).alignment = Alignment(horizontal='center')

            for sheetname_iter in writer.sheets:
                current_sheet = writer.sheets[sheetname_iter]
                for col_idx_iter, column_obj in enumerate(current_sheet.columns):
                    max_cell_length = 0
                    column_letter_val = get_column_letter(col_idx_iter + 1)
                    
                    header_cell_val = current_sheet.cell(row=1, column=col_idx_iter + 1) # Title row
                    if header_cell_val.value and len(str(header_cell_val.value)) > max_cell_length:
                         max_cell_length = len(str(header_cell_val.value))
                    
                    header_cell_val_2 = current_sheet.cell(row=2, column=col_idx_iter + 1) # Header row
                    if header_cell_val_2.value and len(str(header_cell_val_2.value)) > max_cell_length:
                         max_cell_length = len(str(header_cell_val_2.value))

                    for cell_obj in column_obj:
                        if cell_obj.value:
                            cell_len_val = len(str(cell_obj.value))
                            if cell_len_val > max_cell_length:
                                max_cell_length = cell_len_val
                    current_sheet.column_dimensions[column_letter_val].width = max_cell_length + 3 if max_cell_length > 0 else 12
        
        output_buffer.seek(0)
        return output_buffer
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        raise

@app.route('/api/export', methods=['GET'])
@admin_required
def export_data_route():
    """Export data endpoint with enhanced error handling - Admin only"""
    data_type = request.args.get('type')
    start_date_str = request.args.get('startDate')
    end_date_str = request.args.get('endDate')

    if not data_type:
        return jsonify({"error": "Missing 'type' parameter (should be 'items' or 'users')"}), 400
    if data_type not in ['items', 'users']:
        return jsonify({"error": "Invalid 'type' parameter (should be 'items' or 'users')"}), 400

    if not db:
        return jsonify({"error": "Database not available"}), 503

    try:
        excel_buffer = generate_excel_report_to_buffer(
            data_type=data_type, 
            start_date_str=start_date_str, 
            end_date_str=end_date_str
        )
        
        filename_prefix = data_type
        date_range_str = "all_time"
        if start_date_str and end_date_str:
            date_range_str = f"{start_date_str}_to_{end_date_str}"
        elif start_date_str:
            date_range_str = f"from_{start_date_str}"
        elif end_date_str:
            date_range_str = f"up_to_{end_date_str}"
        
        output_filename = f"GCFinder_{filename_prefix}_report_{date_range_str}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_filename 
        )
    except ValueError as ve:
        logger.error(f"ValueError during report generation: {ve}")
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        logger.error(f"Unexpected error during report generation: {e}", exc_info=True)
        return jsonify({"error": "An unexpected error occurred. Please check server logs."}), 500

# ============================================================================
# SECURE API ENDPOINTS - Phase 2 Security Implementation
# ============================================================================

@app.route('/api/users', methods=['GET'])
@admin_required
def get_all_users():
    """Get all users (students + officials) - Admin only"""
    try:
        logger.info("=== Starting user fetch ===")
        logger.info(f"Database connection: {'Connected' if db else 'Not connected'}")
        
        users = []
        doc_count = 0
        
        # Fetch students
        students_ref = db.collection('students')
        logger.info("Getting students collection reference...")
        students_snapshot = students_ref.stream()
        
        for doc in students_snapshot:
            doc_count += 1
            user_data = doc.to_dict()
            users.append({
                'id': doc.id,
                'full_name': user_data.get('full_name', user_data.get('name', 'N/A')),
                'student_id': user_data.get('student_id'),
                'email': user_data.get('email', f"{user_data.get('student_id')}@gordoncollege.edu.ph"),
                'year_level': user_data.get('year_level', 'N/A'),
                'status': user_data.get('status', 'active'),
                'role': user_data.get('role', 'student'),
                'flagReason': user_data.get('flagReason'),
                'flagDuration': user_data.get('flagDuration'),
                'flagExpiresAt': user_data.get('flagExpiresAt'),
                'banReason': user_data.get('banReason'),
                'banDuration': user_data.get('banDuration'),
                'banExpiresAt': user_data.get('banExpiresAt'),
                'profileUrl': user_data.get('profileUrl'),
                'createdAt': user_data.get('createdAt')
            })
        
        logger.info(f"Found {doc_count} students")
        
        # Fetch officials
        officials_ref = db.collection('officials')
        logger.info("Getting officials collection reference...")
        officials_snapshot = officials_ref.stream()
        
        official_count = 0
        for doc in officials_snapshot:
            official_count += 1
            doc_count += 1
            user_data = doc.to_dict()
            users.append({
                'id': doc.id,
                'full_name': user_data.get('full_name', user_data.get('name', 'N/A')),
                'student_id': user_data.get('employee_id'),  # Use employee_id for officials
                'email': user_data.get('email', f"{user_data.get('employee_id')}@gordoncollege.edu.ph"),
                'year_level': 'N/A',  # Officials don't have year levels
                'status': user_data.get('status', 'active'),
                'role': user_data.get('role', 'official'),
                'flagReason': user_data.get('flagReason'),
                'flagDuration': user_data.get('flagDuration'),
                'flagExpiresAt': user_data.get('flagExpiresAt'),
                'banReason': user_data.get('banReason'),
                'banDuration': user_data.get('banDuration'),
                'banExpiresAt': user_data.get('banExpiresAt'),
                'profileUrl': user_data.get('profileUrl'),
                'createdAt': user_data.get('createdAt')
            })
        
        logger.info(f"Found {official_count} officials")
        logger.info(f"=== Fetch complete: Found {doc_count} total documents, processed {len(users)} users ===")
        return jsonify({'users': users}), 200
    except Exception as e:
        logger.error(f"Error fetching users: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch users'}), 500

# ===================== LOST REQUESTS =====================

def _sanitize_text(value, max_len=2000):
    try:
        if value is None:
            return None
        text = str(value).strip()
        if len(text) > max_len:
            text = text[:max_len]
        return text
    except Exception:
        return None

def _get_user_role(uid):
    try:
        user_ref = db.collection('students').document(uid)
        docu = user_ref.get()
        if docu.exists:
            data = docu.to_dict()
            return (data.get('role') or 'student').lower()
    except Exception as e:
        logger.warning(f"Failed to get role for {uid}: {e}")
    return 'student'

# ==== Additional Security Helpers ====
_DATA_URL_IMAGE_RE = re.compile(r'^data:image/(jpeg|jpg|png);base64,[A-Za-z0-9+/=\r\n]+$')
_DATE_YMD_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')

def _is_valid_date_yyyy_mm_dd(date_str: str) -> bool:
    if not isinstance(date_str, str):
        return False
    if not _DATE_YMD_RE.match(date_str):
        return False
    try:
        # Ensure it's a valid calendar date
        datetime.strptime(date_str, '%Y-%m-%d')
        return True
    except Exception:
        return False

def _is_valid_image_data_url(image_url: str) -> bool:
    if not isinstance(image_url, str):
        return False
    if not _DATA_URL_IMAGE_RE.match(image_url.strip()):
        return False
    # Basic size check on base64 length to avoid extremely large payloads (approx < 2MB)
    # 4/3 expansion ratio for base64; be generous: up to ~3MB payloads
    try:
        b64_part = image_url.split(',', 1)[1]
        # Remove newlines for safety
        b64_part = b64_part.replace('\n', '').replace('\r', '')
        # Rough upper bound; skip decoding for performance
        if len(b64_part) > 4_200_000:  # ~3.1MB decoded
            return False
        return True
    except Exception:
        return False

# Simple in-memory rate limiter for lost request creation
_LR_RATE_WINDOW_SECONDS = 60  # 1 minute window
_LR_RATE_MAX_IN_WINDOW = 3    # up to 3 requests/min
_LR_RATE_DAILY_MAX = 30       # up to 30/day
_lost_req_window: dict = {}   # uid -> (window_start_ts, count)
_lost_req_daily: dict = {}    # uid -> (day_ymd_str, count)

def _check_lost_request_rate_limit(uid: str) -> bool:
    now = time.time()
    # Per-minute window
    ws, cnt = _lost_req_window.get(uid, (0, 0))
    if now - ws > _LR_RATE_WINDOW_SECONDS:
        _lost_req_window[uid] = (now, 1)
    else:
        if cnt + 1 > _LR_RATE_MAX_IN_WINDOW:
            return False
        _lost_req_window[uid] = (ws, cnt + 1)
    # Per-day limit
    day_str = datetime.now().strftime('%Y-%m-%d')
    d_day, d_cnt = _lost_req_daily.get(uid, (day_str, 0))
    if d_day != day_str:
        _lost_req_daily[uid] = (day_str, 1)
    else:
        if d_cnt + 1 > _LR_RATE_DAILY_MAX:
            return False
        _lost_req_daily[uid] = (day_str, d_cnt + 1)
    return True

@app.route('/api/lost-requests', methods=['POST'])
@login_required
def create_lost_request():
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    payload = request.get_json() or {}
    item_name = _sanitize_text(payload.get('itemName'), 200)
    description = _sanitize_text(payload.get('description'), 2000)
    date_lost = _sanitize_text(payload.get('dateLost'), 64)
    location_lost = _sanitize_text(payload.get('locationLost'), 256)
    # Allow large data URLs for images (resized client-side). Limit generously to avoid truncation
    image_url = _sanitize_text(payload.get('imageUrl'), 2000000)
    if not all([item_name, description, date_lost, location_lost, image_url]):
        return jsonify({"error": "Missing required fields"}), 400
    # Basic image data URL validation to prevent malformed inputs
    try:
        if not _is_valid_image_data_url(image_url):
            return jsonify({"error": "Invalid image format"}), 400
    except Exception:
        return jsonify({"error": "Invalid image format"}), 400
    # Strict date validation YYYY-MM-DD
    if not _is_valid_date_yyyy_mm_dd(date_lost):
        return jsonify({"error": "Invalid date format (expected YYYY-MM-DD)"}), 400

    uid = request.user['uid']
    # Rate limit to prevent abuse
    if not _check_lost_request_rate_limit(uid):
        return jsonify({"error": "Rate limit exceeded. Please try again later."}), 429

    role = _get_user_role(uid)
    # Fetch requester profile for denormalization
    requester_name = None
    requester_email = None
    try:
        user_doc = db.collection('students').document(uid).get()
        if user_doc.exists:
            u = user_doc.to_dict()
            requester_name = u.get('full_name') or u.get('name')
            requester_email = u.get('email') or (f"{u.get('student_id')}@gordoncollege.edu.ph" if u.get('student_id') else None)
        if not requester_email:
            requester_email = f"{uid}@gordoncollege.edu.ph"
    except Exception as ue:
        logger.warning(f"Could not fetch requester profile for {uid}: {ue}")

    try:
        lost_ref = db.collection('lost_requests')
        doc_ref = lost_ref.document()
        doc_ref.set({
            'itemName': item_name,
            'description': description,
            'dateLost': date_lost,
            'locationLost': location_lost,
            'imageUrl': image_url,
            'requesterId': uid,
            'requesterRole': role,
            'requesterName': requester_name,
            'requesterEmail': requester_email,
            # Store normalized fields as well for convenience/consistency
            'full_name': requester_name,
            'email': requester_email,
            'status': 'pending',
            'createdAt': firestore.SERVER_TIMESTAMP,
            'updatedAt': firestore.SERVER_TIMESTAMP
        })
        return jsonify({'id': doc_ref.id, 'message': 'Lost request submitted', 'status': 'pending'}), 201
    except Exception as e:
        logger.error(f"Error creating lost request: {e}")
        return jsonify({"error": "Failed to submit lost request"}), 500

@app.route('/api/lost-requests', methods=['GET'])
@login_required
def list_lost_requests():
    try:
        uid = request.user['uid']
        # Admin check
        is_admin = db.collection('admin').document(uid).get().exists
        status_filter = request.args.get('status')

        q = db.collection('lost_requests')
        if not is_admin:
            q = q.where('requesterId', '==', uid)
        if status_filter:
            q = q.where('status', '==', status_filter)

        snapshot = q.stream()
        items = []
        for d in snapshot:
            data = d.to_dict()
            created_at = data.get('createdAt')
            updated_at = data.get('updatedAt')
            # Start with any denormalized values stored on the request itself
            requester_name = data.get('requesterName')
            requester_email = data.get('requesterEmail')
            if is_admin:
                try:
                    req_id = data.get('requesterId')
                    if req_id:
                        user_doc = db.collection('students').document(req_id).get()
                        if user_doc.exists:
                            u = user_doc.to_dict()
                            # Prefer live profile data, but fall back to stored values if missing
                            requester_name = (
                                u.get('full_name') or u.get('name') or requester_name or 'N/A'
                            )
                            requester_email = (
                                u.get('email') or (f"{u.get('student_id')}@gordoncollege.edu.ph" if u.get('student_id') else requester_email)
                            )
                except Exception as ue:
                    logger.warning(f"Failed to enrich requester info for {d.id}: {ue}")
            items.append({
                'id': d.id,
                'itemName': data.get('itemName'),
                'description': data.get('description'),
                'dateLost': data.get('dateLost'),
                'locationLost': data.get('locationLost'),
                'imageUrl': data.get('imageUrl'),
                'requesterRole': data.get('requesterRole', 'student'),
                'status': data.get('status', 'pending'),
                **({'requesterName': requester_name, 'requesterEmail': requester_email} if is_admin else {}),
                'createdAt': created_at.isoformat() if hasattr(created_at, 'isoformat') else created_at,
                'updatedAt': updated_at.isoformat() if hasattr(updated_at, 'isoformat') else updated_at
            })
        return jsonify({'requests': items}), 200
    except Exception as e:
        logger.error(f"Error listing lost requests: {e}")
        return jsonify({'error': 'Failed to list lost requests'}), 500

@app.route('/api/lost-requests/<request_id>/approve', methods=['PUT'])
@admin_required
def approve_lost_request(request_id):
    try:
        req_ref = db.collection('lost_requests').document(request_id)
        req_doc = req_ref.get()
        if not req_doc.exists:
            return jsonify({'error': 'Request not found'}), 404

        data = req_doc.to_dict()
        # Create public entry in lost_items (without PII)
        public_ref = db.collection('lost_items').document()
        public_ref.set({
            'itemName': data.get('itemName'),
            'description': data.get('description'),
            'dateLost': data.get('dateLost'),
            'locationLost': data.get('locationLost'),
            'imageUrl': data.get('imageUrl'),
            'postedByRole': data.get('requesterRole', 'student'),
            'status': 'approved',
            'createdAt': firestore.SERVER_TIMESTAMP,
            'sourceRequestId': request_id,
            # Denormalized requester info for admin convenience
            'requesterName': data.get('requesterName'),
            'requesterEmail': data.get('requesterEmail')
        })

        # Update original request
        req_ref.update({
            'status': 'approved',
            'updatedAt': firestore.SERVER_TIMESTAMP
        })

        return jsonify({'message': 'Request approved and published', 'publicId': public_ref.id}), 200
    except Exception as e:
        logger.error(f"Error approving lost request {request_id}: {e}")
        return jsonify({'error': 'Failed to approve lost request'}), 500

@app.route('/api/lost-requests/<request_id>/reject', methods=['PUT'])
@admin_required
def reject_lost_request(request_id):
    try:
        if not request.is_json:
            return jsonify({"error": "Request must be JSON"}), 400
        feedback = _sanitize_text((request.get_json() or {}).get('feedback'), 500)
        req_ref = db.collection('lost_requests').document(request_id)
        req_doc = req_ref.get()
        if not req_doc.exists:
            return jsonify({'error': 'Request not found'}), 404

        update_data = {
            'status': 'rejected',
            'updatedAt': firestore.SERVER_TIMESTAMP
        }
        if feedback:
            update_data['feedback'] = feedback

        req_ref.update(update_data)
        return jsonify({'message': 'Request rejected'}), 200
    except Exception as e:
        logger.error(f"Error rejecting lost request {request_id}: {e}")
        return jsonify({'error': 'Failed to reject lost request'}), 500

@app.route('/api/lost-items', methods=['GET'])
@login_required
def list_lost_items():
    try:
        uid = request.user['uid']
        decoded_email = request.user.get('email')
        is_admin = db.collection('admin').document(uid).get().exists
        # Fallback to profile email if token lacks it
        viewer_email = decoded_email
        if not viewer_email:
            try:
                user_doc = db.collection('students').document(uid).get()
                if user_doc.exists:
                    u = user_doc.to_dict()
                    viewer_email = u.get('email') or (f"{u.get('student_id')}@gordoncollege.edu.ph" if u.get('student_id') else None)
            except Exception as ue:
                logger.warning(f"Could not resolve viewer email for {uid}: {ue}")
        q = db.collection('lost_items').stream()
        results = []
        for d in q:
            data = d.to_dict()
            created_at = data.get('createdAt')
            # Hide resolved/archived items from non-admin viewers
            item_status = (data.get('status') or 'approved').lower()
            if not is_admin and item_status in ('resolved', 'archived'):
                continue
            # Build requester info for admins, compute ownership flag for all
            requester_name = None
            requester_email = None
            if is_admin:
                try:
                    source_id = data.get('sourceRequestId')
                    if source_id:
                        req_doc = db.collection('lost_requests').document(source_id).get()
                        if req_doc.exists:
                            req_data = req_doc.to_dict()
                            rid = req_data.get('requesterId')
                            if rid:
                                user_doc = db.collection('students').document(rid).get()
                                if user_doc.exists:
                                    u = user_doc.to_dict()
                                    requester_name = u.get('full_name') or u.get('name') or 'N/A'
                                    requester_email = u.get('email') or f"{u.get('student_id')}@gordoncollege.edu.ph"
                except Exception as ue:
                    logger.warning(f"Failed to enrich lost item {d.id} with requester info: {ue}")
            # Ownership flag (no PII exposure)
            is_owner = False
            try:
                # Prefer direct email compare if available on item
                item_email = data.get('requesterEmail')
                if not item_email:
                    # fallback to source request's stored email or uid
                    source_id = data.get('sourceRequestId')
                    if source_id:
                        req_doc = db.collection('lost_requests').document(source_id).get()
                        if req_doc.exists:
                            req_data = req_doc.to_dict()
                            item_email = req_data.get('requesterEmail')
                            if not item_email and req_data.get('requesterId') == uid:
                                is_owner = True
                if not is_owner and item_email and viewer_email:
                    is_owner = (item_email.lower() == viewer_email.lower())
            except Exception as oe:
                logger.warning(f"Ownership compute failed for lost item {d.id}: {oe}")
            results.append({
                'id': d.id,
                'itemName': data.get('itemName'),
                'description': data.get('description'),
                'dateLost': data.get('dateLost'),
                'locationLost': data.get('locationLost'),
                'imageUrl': data.get('imageUrl'),
                'postedByRole': data.get('postedByRole', 'student'),
                'isOwner': is_owner,
                **({'requesterName': requester_name, 'requesterEmail': requester_email, 'status': item_status} if is_admin else {}),
                'createdAt': created_at.isoformat() if hasattr(created_at, 'isoformat') else created_at
            })
        return jsonify({'items': results}), 200
    except Exception as e:
        logger.error(f"Error listing lost items: {e}")
        return jsonify({'error': 'Failed to list lost items'}), 500

@app.route('/api/lost-items/<item_id>/resolve', methods=['PUT'])
@admin_required
def resolve_lost_item(item_id):
    try:
        ref = db.collection('lost_items').document(item_id)
        doc = ref.get()
        if not doc.exists:
            return jsonify({'error': 'Lost item not found'}), 404
        current = (doc.to_dict() or {}).get('status', 'approved').lower()
        # Only allow resolve from approved
        if current not in ('approved',):
            return jsonify({'error': f'Invalid transition from {current} to resolved'}), 400
        ref.update({
            'status': 'resolved',
            'updatedAt': firestore.SERVER_TIMESTAMP
        })
        return jsonify({'message': 'Lost item marked as resolved'}), 200
    except Exception as e:
        logger.error(f"Error resolving lost item {item_id}: {e}")
        return jsonify({'error': 'Failed to resolve lost item'}), 500

@app.route('/api/lost-items/<item_id>/unresolve', methods=['PUT'])
@admin_required
def unresolve_lost_item(item_id):
    try:
        ref = db.collection('lost_items').document(item_id)
        doc = ref.get()
        if not doc.exists:
            return jsonify({'error': 'Lost item not found'}), 404
        current = (doc.to_dict() or {}).get('status', 'approved').lower()
        # Only allow unresolve from resolved
        if current not in ('resolved',):
            return jsonify({'error': f'Invalid transition from {current} to approved'}), 400
        # Revert back to approved so it shows publicly again
        ref.update({
            'status': 'approved',
            'updatedAt': firestore.SERVER_TIMESTAMP
        })
        return jsonify({'message': 'Lost item un-resolved'}), 200
    except Exception as e:
        logger.error(f"Error unresolving lost item {item_id}: {e}")
        return jsonify({'error': 'Failed to unresolve lost item'}), 500

@app.route('/api/lost-items/<item_id>/archive', methods=['PUT'])
@admin_required
def archive_lost_item(item_id):
    try:
        ref = db.collection('lost_items').document(item_id)
        doc = ref.get()
        if not doc.exists:
            return jsonify({'error': 'Lost item not found'}), 404
        current = (doc.to_dict() or {}).get('status', 'approved').lower()
        # Allow archive from approved or resolved
        if current not in ('approved', 'resolved'):
            return jsonify({'error': f'Invalid transition from {current} to archived'}), 400
        ref.update({
            'status': 'Archived',
            'updatedAt': firestore.SERVER_TIMESTAMP
        })
        return jsonify({'message': 'Lost item archived'}), 200
    except Exception as e:
        logger.error(f"Error archiving lost item {item_id}: {e}")
        return jsonify({'error': 'Failed to archive lost item'}), 500

@app.route('/api/lost-items/<item_id>/unarchive', methods=['PUT'])
@admin_required
def unarchive_lost_item(item_id):
    try:
        ref = db.collection('lost_items').document(item_id)
        doc = ref.get()
        if not doc.exists:
            return jsonify({'error': 'Lost item not found'}), 404
        current = (doc.to_dict() or {}).get('status', 'approved').lower()
        # Only allow unarchive from archived
        if current not in ('archived',):
            return jsonify({'error': f'Invalid transition from {current} to approved'}), 400
        # Revert back to approved so it shows publicly again
        ref.update({
            'status': 'approved',
            'updatedAt': firestore.SERVER_TIMESTAMP
        })
        return jsonify({'message': 'Lost item un-archived'}), 200
    except Exception as e:
        logger.error(f"Error unarchiving lost item {item_id}: {e}")
        return jsonify({'error': 'Failed to unarchive lost item'}), 500

@app.route('/api/users/<uid>/status', methods=['PUT'])
@admin_required
def update_user_status(uid):
    """Update user status (flag/ban/unban) - Admin only. Checks both students and officials collections."""
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400
    
    status_data = request.get_json()
    
    try:
        # Check students collection first
        user_ref = db.collection('students').document(uid)
        user_doc = user_ref.get()
        
        # If not in students, check officials
        if not user_doc.exists:
            user_ref = db.collection('officials').document(uid)
            user_doc = user_ref.get()
        
        if not user_doc.exists:
            return jsonify({"error": "User not found"}), 404
        
        # Build update data
        update_data = {
            'status': status_data.get('status'),
            'updatedAt': firestore.SERVER_TIMESTAMP
        }
        
        # Add flag/ban specific fields
        if 'flagReason' in status_data:
            update_data['flagReason'] = status_data['flagReason']
        if 'flagDuration' in status_data:
            update_data['flagDuration'] = status_data['flagDuration']
        if 'flagExpiresAt' in status_data:
            update_data['flagExpiresAt'] = status_data['flagExpiresAt']
        if 'banReason' in status_data:
            update_data['banReason'] = status_data['banReason']
        if 'banDuration' in status_data:
            update_data['banDuration'] = status_data['banDuration']
        if 'banExpiresAt' in status_data:
            update_data['banExpiresAt'] = status_data['banExpiresAt']
        
        user_ref.update(update_data)
        
        return jsonify({
            "message": f"User {uid} status updated successfully",
            "status": status_data.get('status')
        }), 200
        
    except Exception as e:
        logger.error(f"Error updating user status: {e}")
        return jsonify({"error": "Failed to update user status"}), 500

@app.route('/api/items', methods=['GET'])
@login_required
def get_items():
    """Get items with role-based filtering"""
    try:
        logger.info("=== Fetching items ===")
        uid = request.user['uid']
        logger.info(f"User ID: {uid}")
        
        # Check if user is admin
        admin_ref = db.collection('admin').document(uid)
        is_admin = admin_ref.get().exists
        logger.info(f"Is admin: {is_admin}")
        
        items_ref = db.collection('items')
        
        # Admins see all items, students see only approved items
        if is_admin:
            logger.info("Fetching ALL items for admin")
            items_snapshot = items_ref.stream()
        else:
            logger.info("Fetching APPROVED items for student")
            items_snapshot = items_ref.where('adminApproval', '==', True).stream()
        
        items = []
        doc_count = 0
        for doc in items_snapshot:
            doc_count += 1
            logger.info(f"Processing document {doc_count}: {doc.id}")
            try:
                item_data = doc.to_dict()
                
                # Convert Firestore Timestamps to ISO format strings for JSON serialization
                created_at = item_data.get('createdAt')
                updated_at = item_data.get('updatedAt')
                date_field = item_data.get('date')
                
                items.append({
                    'id': doc.id,
                    'name': item_data.get('name'),
                    'category': item_data.get('category'),
                    'description': item_data.get('description'),
                    'location': item_data.get('location'),
                    'exactLocation': item_data.get('exactLocation'),
                    'date': date_field.isoformat() if hasattr(date_field, 'isoformat') else date_field,
                    'status': item_data.get('status'),
                    'adminApproval': item_data.get('adminApproval'),
                    'imageData': item_data.get('imageData'),
                    'submitter': item_data.get('submitter'),
                    'uniqueIdentifier': item_data.get('uniqueIdentifier'),
                    'additionalDetails': item_data.get('additionalDetails'),
                    'createdAt': created_at.isoformat() if hasattr(created_at, 'isoformat') else created_at,
                    'updatedAt': updated_at.isoformat() if hasattr(updated_at, 'isoformat') else updated_at
                })
            except Exception as doc_error:
                logger.error(f"Error processing document {doc.id}: {doc_error}", exc_info=True)
                continue
        
        logger.info(f"=== Successfully fetched {len(items)} items out of {doc_count} documents ===")
        return jsonify({'items': items}), 200
    except Exception as e:
        logger.error(f"Error fetching items: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch items'}), 500

@app.route('/api/items/pending', methods=['GET'])
@admin_required
def get_pending_items():
    """Get items pending admin approval - Admin only"""
    try:
        items_ref = db.collection('items')
        items_snapshot = items_ref.where('adminApproval', '==', False).stream()
        
        items = []
        for doc in items_snapshot:
            item_data = doc.to_dict()
            
            # Convert Firestore Timestamps to ISO format strings for JSON serialization
            created_at = item_data.get('createdAt')
            updated_at = item_data.get('updatedAt')
            date_field = item_data.get('date')
            
            items.append({
                'id': doc.id,
                'title': item_data.get('name'),
                'name': item_data.get('name'),
                'category': item_data.get('category'),
                'status': 'pending',
                'date': date_field.isoformat() if hasattr(date_field, 'isoformat') else date_field,
                'image': item_data.get('imageData', [])[0].get('dataUrl') if item_data.get('imageData') and len(item_data.get('imageData')) > 0 else None,
                'images': [img.get('dataUrl') for img in item_data.get('imageData', [])] if item_data.get('imageData') else [],
                'description': item_data.get('description'),
                'location': item_data.get('location'),
                'exactLocation': item_data.get('exactLocation'),
                'uniqueIdentifier': item_data.get('uniqueIdentifier'),
                'additionalDetails': item_data.get('additionalDetails'),
                'submitter': {
                    'full_name': item_data.get('submitter', {}).get('full_name', 'N/A'),
                    'student_id': item_data.get('submitter', {}).get('student_id')
                },
                'createdAt': created_at.isoformat() if hasattr(created_at, 'isoformat') else created_at,
                'updatedAt': updated_at.isoformat() if hasattr(updated_at, 'isoformat') else updated_at,
                'adminApproval': item_data.get('adminApproval')
            })
        
        return jsonify({'items': items}), 200
    except Exception as e:
        logger.error(f"Error fetching pending items: {e}")
        return jsonify({'error': 'Failed to fetch pending items'}), 500

@app.route('/api/items/browse', methods=['GET'])
@login_required
def browse_items():
    """Get items for student browsing - excludes claimed items and applies visibility rules"""
    try:
        uid = request.user['uid']
        
        # Get user's claimed item IDs
        claims_ref = db.collection('claims')
        user_claims = claims_ref.where('claimerId', '==', uid).stream()
        claimed_item_ids = set()
        for claim_doc in user_claims:
            claim_data = claim_doc.to_dict()
            claimed_item_ids.add(claim_data.get('itemId'))
        
        # Get all non-archived items
        items_ref = db.collection('items')
        items_snapshot = items_ref.where('status', '!=', 'Archived').stream()
        
        items = []
        user_submitted_count = 0
        
        for doc in items_snapshot:
            item_data = doc.to_dict()
            item_id = doc.id
            
            # Check if user is the submitter
            submitter = item_data.get('submitter', {})
            is_submitter = submitter.get('student_id') == uid if submitter else False
            
            if is_submitter:
                user_submitted_count += 1
            
            # Determine visibility
            status = item_data.get('status')
            admin_approval = item_data.get('adminApproval', False)
            is_disapproved = status == 'Disapproved'
            
            if is_disapproved:
                is_visible = is_submitter
            else:
                # Visible if admin approved OR if user is the submitter
                is_visible = admin_approval or is_submitter
            
            # Include if visible and not already claimed by this user
            if is_visible and item_id not in claimed_item_ids:
                # Convert timestamps
                created_at = item_data.get('createdAt')
                updated_at = item_data.get('updatedAt')
                date_field = item_data.get('date')
                
                items.append({
                    'id': item_id,
                    'name': item_data.get('name'),
                    'category': item_data.get('category'),
                    'location': item_data.get('location'),
                    'date': date_field.isoformat() if hasattr(date_field, 'isoformat') else date_field,
                    'status': item_data.get('status'),
                    'description': item_data.get('description'),
                    'imageData': item_data.get('imageData'),
                    'image': item_data.get('imageData', [])[0].get('dataUrl') if item_data.get('imageData') and len(item_data.get('imageData')) > 0 else None,
                    'exactLocation': item_data.get('exactLocation'),
                    'uniqueIdentifier': item_data.get('uniqueIdentifier'),
                    'additionalDetails': item_data.get('additionalDetails'),
                    'submitter': item_data.get('submitter'),
                    'adminApproval': admin_approval,
                    'createdAt': created_at.isoformat() if hasattr(created_at, 'isoformat') else created_at,
                    'updatedAt': updated_at.isoformat() if hasattr(updated_at, 'isoformat') else updated_at
                })
        
        return jsonify({
            'items': items,
            'userSubmittedCount': user_submitted_count
        }), 200
        
    except Exception as e:
        logger.error(f"Error browsing items: {e}", exc_info=True)
        return jsonify({'error': 'Failed to browse items'}), 500

@app.route('/api/items/<item_id>', methods=['PUT'])
@admin_required
def update_item(item_id):
    """Update item - Admin only"""
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400
    
    update_data = request.get_json()
    
    try:
        item_ref = db.collection('items').document(item_id)
        item_doc = item_ref.get()
        
        if not item_doc.exists:
            return jsonify({"error": "Item not found"}), 404
        
        # Add timestamp
        update_data['updatedAt'] = firestore.SERVER_TIMESTAMP
        
        item_ref.update(update_data)
        
        return jsonify({
            "message": f"Item {item_id} updated successfully"
        }), 200
        
    except Exception as e:
        logger.error(f"Error updating item: {e}")
        return jsonify({"error": "Failed to update item"}), 500

@app.route('/api/items/<item_id>/approve', methods=['POST'])
@admin_required
def approve_item(item_id):
    """Approve a pending item - Admin only"""
    try:
        item_ref = db.collection('items').document(item_id)
        item_doc = item_ref.get()
        
        if not item_doc.exists:
            return jsonify({"error": "Item not found"}), 404
        
        item_ref.update({
            'adminApproval': True,
            'status': 'Unclaimed',
            'updatedAt': firestore.SERVER_TIMESTAMP
        })
        
        return jsonify({"message": f"Item {item_id} approved successfully"}), 200
        
    except Exception as e:
        logger.error(f"Error approving item: {e}")
        return jsonify({"error": "Failed to approve item"}), 500

@app.route('/api/items/<item_id>/disapprove', methods=['POST'])
@admin_required
def disapprove_item(item_id):
    """Disapprove a pending item - Admin only"""
    try:
        item_ref = db.collection('items').document(item_id)
        item_doc = item_ref.get()
        
        if not item_doc.exists:
            return jsonify({"error": "Item not found"}), 404
        
        item_ref.update({
            'status': 'Disapproved',
            'adminApproval': True,  # Set to true so it doesn't show in pending
            'updatedAt': firestore.SERVER_TIMESTAMP
        })
        
        return jsonify({"message": f"Item {item_id} disapproved successfully"}), 200
        
    except Exception as e:
        logger.error(f"Error disapproving item: {e}")
        return jsonify({"error": "Failed to disapprove item"}), 500

@app.route('/api/items/<item_id>', methods=['DELETE'])
@admin_required
def delete_item_endpoint(item_id):
    """Delete item and associated claims - Admin only"""
    try:
        # Delete item
        item_ref = db.collection('items').document(item_id)
        item_doc = item_ref.get()
        
        if not item_doc.exists:
            return jsonify({"error": "Item not found"}), 404
        
        item_ref.delete()
        
        # Delete associated claims
        claims_ref = db.collection('claims')
        claims_query = claims_ref.where('itemId', '==', item_id)
        claims_snapshot = claims_query.stream()
        
        deleted_claims = 0
        for claim_doc in claims_snapshot:
            claim_doc.reference.delete()
            deleted_claims += 1
        
        return jsonify({
            "message": f"Item {item_id} and {deleted_claims} associated claims deleted successfully"
        }), 200
        
    except Exception as e:
        logger.error(f"Error deleting item: {e}")
        return jsonify({"error": "Failed to delete item"}), 500

@app.route('/api/items/<item_id>/archive', methods=['POST'])
@admin_required
def archive_item(item_id):
    """Archive item - Admin only"""
    try:
        item_ref = db.collection('items').document(item_id)
        item_doc = item_ref.get()
        
        if not item_doc.exists:
            return jsonify({"error": "Item not found"}), 404
        
        item_data = item_doc.to_dict()
        current_status = item_data.get('status')
        
        item_ref.update({
            'status': 'Archived',
            'previousStatus': current_status,
            'updatedAt': firestore.SERVER_TIMESTAMP
        })
        
        # Delete pending claims for archived item
        claims_ref = db.collection('claims')
        pending_claims = claims_ref.where('itemId', '==', item_id).where('claimStatus', '==', 'Pending').stream()
        
        deleted_claims = 0
        for claim_doc in pending_claims:
            claim_doc.reference.delete()
            deleted_claims += 1
        
        return jsonify({
            "message": f"Item {item_id} archived successfully",
            "deleted_pending_claims": deleted_claims
        }), 200
        
    except Exception as e:
        logger.error(f"Error archiving item: {e}")
        return jsonify({"error": "Failed to archive item"}), 500

@app.route('/api/items/<item_id>/unarchive', methods=['POST'])
@admin_required
def unarchive_item(item_id):
    """Unarchive item - Admin only"""
    try:
        item_ref = db.collection('items').document(item_id)
        item_doc = item_ref.get()
        
        if not item_doc.exists:
            return jsonify({"error": "Item not found"}), 404
        
        item_data = item_doc.to_dict()
        previous_status = item_data.get('previousStatus', 'Unclaimed')
        
        item_ref.update({
            'status': previous_status,
            'previousStatus': firestore.DELETE_FIELD,
            'updatedAt': firestore.SERVER_TIMESTAMP
        })
        
        return jsonify({
            "message": f"Item {item_id} unarchived successfully",
            "restored_status": previous_status
        }), 200
        
    except Exception as e:
        logger.error(f"Error unarchiving item: {e}")
        return jsonify({"error": "Failed to unarchive item"}), 500

@app.route('/api/claims', methods=['GET'])
@login_required
def get_claims():
    """Get claims with role-based filtering"""
    try:
        uid = request.user['uid']
        
        # Check if user is admin
        admin_ref = db.collection('admin').document(uid)
        is_admin = admin_ref.get().exists
        
        claims_ref = db.collection('claims')
        
        # Admins see all claims, students see only their own
        if is_admin:
            claims_snapshot = claims_ref.stream()
        else:
            claims_snapshot = claims_ref.where('claimerId', '==', uid).stream()
        
        claims = []
        for doc in claims_snapshot:
            claim_data = doc.to_dict()
            
            # Convert Firestore Timestamps to ISO format strings for JSON serialization
            created_at = claim_data.get('createdAt')
            updated_at = claim_data.get('updatedAt')
            
            # Fetch the item's image from the items collection
            item_image = None
            item_id = claim_data.get('itemId')
            if item_id:
                try:
                    item_ref = db.collection('items').document(item_id)
                    item_doc = item_ref.get()
                    if item_doc.exists:
                        item_data = item_doc.to_dict()
                        image_data = item_data.get('imageData')
                        if image_data and len(image_data) > 0:
                            item_image = image_data[0].get('dataUrl')
                except Exception as item_error:
                    logger.warning(f"Could not fetch item image for claim {doc.id}: {item_error}")
            
            claims.append({
                'id': doc.id,
                'itemId': item_id,
                'itemName': claim_data.get('itemName'),
                'itemImage': item_image,
                'claimerId': claim_data.get('claimerId'),
                'claimerName': claim_data.get('claimerName'),
                'claimStatus': claim_data.get('claimStatus'),
                'lastSeenLocation': claim_data.get('lastSeenLocation'),
                'uniqueIdentifier': claim_data.get('uniqueIdentifier'),
                'additionalDetails': claim_data.get('additionalDetails'),
                'proofImageUrl': claim_data.get('proofImageUrl'),
                'rejectionReason': claim_data.get('rejectionReason'),
                'ownershipProof': claim_data.get('ownershipProof'),
                'createdAt': created_at.isoformat() if hasattr(created_at, 'isoformat') else created_at,
                'updatedAt': updated_at.isoformat() if hasattr(updated_at, 'isoformat') else updated_at
            })
        
        return jsonify({'claims': claims}), 200
    except Exception as e:
        logger.error(f"Error fetching claims: {e}")
        return jsonify({'error': 'Failed to fetch claims'}), 500

@app.route('/api/claims/<claim_id>', methods=['PUT'])
@admin_required
def update_claim(claim_id):
    """Update claim status - Admin only"""
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400
    
    update_data = request.get_json()
    
    try:
        claim_ref = db.collection('claims').document(claim_id)
        claim_doc = claim_ref.get()
        
        if not claim_doc.exists:
            return jsonify({"error": "Claim not found"}), 404
        
        # Add timestamp
        update_data['updatedAt'] = firestore.SERVER_TIMESTAMP
        
        claim_ref.update(update_data)
        
        # If claim is approved/rejected, update related item status
        if 'claimStatus' in update_data and 'itemId' in claim_doc.to_dict():
            item_id = claim_doc.to_dict()['itemId']
            item_ref = db.collection('items').document(item_id)
            
            if update_data['claimStatus'] == 'Approved':
                item_ref.update({
                    'status': 'Claiming',
                    'updatedAt': firestore.SERVER_TIMESTAMP
                })
            elif update_data['claimStatus'] == 'Rejected':
                item_ref.update({
                    'status': 'Unclaimed',
                    'updatedAt': firestore.SERVER_TIMESTAMP
                })
            elif update_data['claimStatus'] == 'Claimed':
                item_ref.update({
                    'status': 'Claimed',
                    'updatedAt': firestore.SERVER_TIMESTAMP
                })
        
        return jsonify({
            "message": f"Claim {claim_id} updated successfully"
        }), 200
        
    except Exception as e:
        logger.error(f"Error updating claim: {e}")
        return jsonify({"error": "Failed to update claim"}), 500

@app.route('/api/dashboard/stats', methods=['GET'])
@admin_required
def get_dashboard_stats():
    """Get dashboard statistics - Admin only"""
    try:
        from datetime import datetime, timedelta
        
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday = today - timedelta(days=1)
        
        # Get pending items
        items_ref = db.collection('items')
        pending_items_snapshot = items_ref.where('adminApproval', '==', False).stream()
        pending_reports_count = 0
        pending_reports_today_yesterday = 0
        
        for doc in pending_items_snapshot:
            pending_reports_count += 1
            item_data = doc.to_dict()
            created_at = item_data.get('createdAt')
            if created_at and hasattr(created_at, 'timestamp'):
                created_date = datetime.fromtimestamp(created_at.timestamp()).replace(hour=0, minute=0, second=0, microsecond=0)
                if created_date >= yesterday:
                    pending_reports_today_yesterday += 1
        
        # Get pending claims
        claims_ref = db.collection('claims')
        pending_claims_snapshot = claims_ref.where('claimStatus', '==', 'Pending').stream()
        pending_claims_count = 0
        pending_claims_today_yesterday = 0
        
        for doc in pending_claims_snapshot:
            pending_claims_count += 1
            claim_data = doc.to_dict()
            created_at = claim_data.get('createdAt')
            if created_at and hasattr(created_at, 'timestamp'):
                created_date = datetime.fromtimestamp(created_at.timestamp()).replace(hour=0, minute=0, second=0, microsecond=0)
                if created_date >= yesterday:
                    pending_claims_today_yesterday += 1
        
        # Get active items
        active_items_snapshot = items_ref.where('status', '==', 'Unclaimed').stream()
        active_items_count = 0
        active_items_today_yesterday = 0
        
        for doc in active_items_snapshot:
            active_items_count += 1
            item_data = doc.to_dict()
            updated_at = item_data.get('updatedAt')
            if updated_at and hasattr(updated_at, 'timestamp'):
                updated_date = datetime.fromtimestamp(updated_at.timestamp()).replace(hour=0, minute=0, second=0, microsecond=0)
                if updated_date >= yesterday:
                    active_items_today_yesterday += 1
        
        # Get item category distribution
        all_items_snapshot = items_ref.stream()
        category_counts = {}
        resolved_count = 0
        unresolved_count = 0
        
        for doc in all_items_snapshot:
            item_data = doc.to_dict()
            
            # Category distribution
            category = item_data.get('category', 'Uncategorized')
            category_counts[category] = category_counts.get(category, 0) + 1
            
            # Resolution rate
            status = item_data.get('status')
            admin_approval = item_data.get('adminApproval', False)
            
            if status in ['Claimed', 'Claiming']:
                resolved_count += 1
            elif status == 'Unclaimed' or (status == 'Pending' and admin_approval) or status == 'Archived':
                unresolved_count += 1
        
        # Format category distribution for pie chart
        item_category_distribution = [{'name': key, 'value': value} for key, value in category_counts.items()]
        
        # Format resolution data for pie chart
        report_resolution_data = [
            {'name': 'Resolved', 'value': resolved_count},
            {'name': 'Unresolved', 'value': unresolved_count}
        ]
        
        stats = {
            'pendingReports': {
                'count': pending_reports_count,
                'fromYesterday': pending_reports_today_yesterday
            },
            'pendingClaims': {
                'count': pending_claims_count,
                'fromYesterday': pending_claims_today_yesterday
            },
            'activeItems': {
                'count': active_items_count,
                'fromYesterday': active_items_today_yesterday
            },
            'itemCategoryDistribution': item_category_distribution,
            'reportResolutionData': report_resolution_data
        }
        
        return jsonify({'stats': stats}), 200
        
    except Exception as e:
        logger.error(f"Error fetching dashboard stats: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch dashboard statistics'}), 500

# ======= CLEANUP / CRON ENDPOINTS =======

@app.route('/api/cron/cleanup-items', methods=['POST', 'GET'])
def cleanup_old_items():
    """
    Cleanup endpoint for automated maintenance:
    1. Archive items with status 'unclaimed', 'pending', 'claimed', 'disapproved' that are older than 15 days
    2. Delete lost requests that are older than 15 days
    
    Can be triggered by cron job with proper authentication.
    Requires CRON_SECRET environment variable to be set for security.
    """
    # Check for cron secret (required for all access)
    cron_secret = request.headers.get('X-Cron-Secret') or request.args.get('secret')
    expected_secret = os.environ.get('CRON_SECRET', '')
    
    # SECURITY: Always require a valid secret - no bypass allowed
    if not expected_secret:
        logger.warning("CRON_SECRET not configured - cleanup endpoint disabled for security")
        return jsonify({'error': 'Cleanup endpoint not configured. Set CRON_SECRET environment variable.'}), 503
    
    if cron_secret != expected_secret:
        logger.warning(f"Unauthorized cleanup attempt from {request.remote_addr}")
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        cutoff_date = datetime.now() - timedelta(days=15)
        results = {
            'items_archived': 0,
            'lost_requests_deleted': 0,
            'errors': []
        }
        
        # 1. Archive old items with specified statuses
        # Include both capitalized and lowercase versions for compatibility
        statuses_to_archive = ['Unclaimed', 'unclaimed', 'Pending', 'pending', 'Claimed', 'claimed', 'Disapproved', 'disapproved']
        items_ref = db.collection('items')
        
        for status in statuses_to_archive:
            try:
                # Query items with this status
                query = items_ref.where('status', '==', status).stream()
                
                for item_doc in query:
                    item_data = item_doc.to_dict()
                    
                    # Get the item's date (dateFound or createdAt)
                    item_date = None
                    date_field = item_data.get('date') or item_data.get('dateFound') or item_data.get('createdAt')
                    
                    if date_field:
                        if hasattr(date_field, 'timestamp'):
                            # Firestore Timestamp
                            item_date = datetime.fromtimestamp(date_field.timestamp())
                        elif isinstance(date_field, datetime):
                            item_date = date_field
                        elif isinstance(date_field, str):
                            # Try parsing string date
                            for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%dT%H:%M:%SZ'):
                                try:
                                    item_date = datetime.strptime(date_field, fmt)
                                    break
                                except ValueError:
                                    continue
                    
                    # If we have a valid date and it's older than 15 days, archive it
                    if item_date and item_date < cutoff_date:
                        try:
                            # Store previous status before archiving
                            item_doc.reference.update({
                                'previousStatus': status,
                                'status': 'Archived',
                                'archivedAt': firestore.SERVER_TIMESTAMP,
                                'archivedReason': 'auto-cleanup-15-days',
                                'updatedAt': firestore.SERVER_TIMESTAMP
                            })
                            results['items_archived'] += 1
                            logger.info(f"Auto-archived item {item_doc.id} (status: {status}, date: {item_date})")
                        except Exception as e:
                            results['errors'].append(f"Failed to archive item {item_doc.id}: {str(e)}")
                            
            except Exception as e:
                results['errors'].append(f"Error processing status '{status}': {str(e)}")
        
        # 2. Delete old lost requests (both pending and approved that are old)
        try:
            lost_requests_ref = db.collection('lost_requests')
            lost_requests = lost_requests_ref.stream()
            
            for req_doc in lost_requests:
                req_data = req_doc.to_dict()
                
                # Get request date
                req_date = None
                date_field = req_data.get('createdAt') or req_data.get('dateLost')
                
                if date_field:
                    if hasattr(date_field, 'timestamp'):
                        req_date = datetime.fromtimestamp(date_field.timestamp())
                    elif isinstance(date_field, datetime):
                        req_date = date_field
                    elif isinstance(date_field, str):
                        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y-%m-%dT%H:%M:%S.%fZ'):
                            try:
                                req_date = datetime.strptime(date_field, fmt)
                                break
                            except ValueError:
                                continue
                
                if req_date and req_date < cutoff_date:
                    try:
                        req_doc.reference.delete()
                        results['lost_requests_deleted'] += 1
                        logger.info(f"Auto-deleted lost request {req_doc.id} (date: {req_date})")
                    except Exception as e:
                        results['errors'].append(f"Failed to delete lost request {req_doc.id}: {str(e)}")
                        
        except Exception as e:
            results['errors'].append(f"Error processing lost requests: {str(e)}")
        
        # 3. Also delete old lost_items (the published/approved ones)
        try:
            lost_items_ref = db.collection('lost_items')
            lost_items = lost_items_ref.stream()
            
            for item_doc in lost_items:
                item_data = item_doc.to_dict()
                
                item_date = None
                date_field = item_data.get('createdAt') or item_data.get('dateLost')
                
                if date_field:
                    if hasattr(date_field, 'timestamp'):
                        item_date = datetime.fromtimestamp(date_field.timestamp())
                    elif isinstance(item_data, datetime):
                        item_date = date_field
                    elif isinstance(date_field, str):
                        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y-%m-%dT%H:%M:%S.%fZ'):
                            try:
                                item_date = datetime.strptime(date_field, fmt)
                                break
                            except ValueError:
                                continue
                
                if item_date and item_date < cutoff_date:
                    try:
                        item_doc.reference.delete()
                        results['lost_requests_deleted'] += 1
                        logger.info(f"Auto-deleted lost item {item_doc.id} (date: {item_date})")
                    except Exception as e:
                        results['errors'].append(f"Failed to delete lost item {item_doc.id}: {str(e)}")
                        
        except Exception as e:
            results['errors'].append(f"Error processing lost items: {str(e)}")
        
        logger.info(f"Cleanup completed: {results['items_archived']} items archived, {results['lost_requests_deleted']} lost requests/items deleted")
        
        return jsonify({
            'success': True,
            'message': 'Cleanup completed',
            'results': results
        }), 200
        
    except Exception as e:
        logger.error(f"Cleanup cron error: {e}", exc_info=True)
        return jsonify({'error': f'Cleanup failed: {str(e)}'}), 500


@app.route('/api/items/archived/delete-all', methods=['DELETE'])
@admin_required
def delete_all_archived_items():
    """
    Delete all archived items permanently - Admin only
    """
    try:
        items_ref = db.collection('items')
        # Check both capitalized and lowercase for compatibility
        archived_items_cap = list(items_ref.where('status', '==', 'Archived').stream())
        archived_items_low = list(items_ref.where('status', '==', 'archived').stream())
        archived_items = archived_items_cap + archived_items_low
        
        deleted_count = 0
        errors = []
        
        for item_doc in archived_items:
            try:
                item_id = item_doc.id
                
                # Also delete associated claims
                claims_ref = db.collection('claims')
                claims_query = claims_ref.where('itemId', '==', item_id).stream()
                
                for claim_doc in claims_query:
                    claim_doc.reference.delete()
                
                # Delete the item
                item_doc.reference.delete()
                deleted_count += 1
                
            except Exception as e:
                errors.append(f"Failed to delete item {item_doc.id}: {str(e)}")
        
        logger.info(f"Deleted {deleted_count} archived items")
        
        return jsonify({
            'success': True,
            'message': f'Successfully deleted {deleted_count} archived items',
            'deleted_count': deleted_count,
            'errors': errors if errors else None
        }), 200
        
    except Exception as e:
        logger.error(f"Error deleting all archived items: {e}", exc_info=True)
        return jsonify({'error': 'Failed to delete archived items'}), 500

# Production-ready error handlers
@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    logger.error(f"Internal server error: {error}")
    return jsonify({'error': 'Internal server error'}), 500

@app.errorhandler(413)
def too_large(error):
    return jsonify({'error': 'File too large'}), 413

if __name__ == '__main__':
    # Production-ready configuration
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('DEBUG', 'False').lower() in ['true', '1', 'yes']
    
    if not db:
        logger.error("Cannot start server: Database connection failed")
        sys.exit(1)
    
    logger.info(f"Starting GCFinder API server on port {port}")
    logger.info(f"Debug mode: {debug}")
    logger.info(f"Database: {'Connected' if db else 'Disconnected'}")
    logger.info(f"HuggingFace Token: {'Set' if HF_API_TOKEN else 'Not Set'}")
    
    app.run(host='0.0.0.0', port=port, debug=debug)
