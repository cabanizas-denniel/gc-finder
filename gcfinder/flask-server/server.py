import os
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import torch
from PIL import Image
from transformers import CLIPProcessor, CLIPModel
import numpy as np
import firebase_admin
from firebase_admin import credentials, firestore
import base64
from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

# Initialize Firebase with credentials from environment variable
firebase_credentials = json.loads(os.getenv('FIREBASE_CREDENTIALS', '{}'))
cred = credentials.Certificate(firebase_credentials)
firebase_admin.initialize_app(cred)
db = firestore.client()

app = Flask(__name__)
allowed_origins = os.getenv('ALLOWED_ORIGINS', 'http://localhost:3000').split(',')
CORS(app, resources={
    r"/api/*": {
        "origins": allowed_origins,
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

# Initialize CLIP
device = "cuda" if torch.cuda.is_available() else "cpu"
model = CLIPModel.from_pretrained("openai/clip-vit-base-patch32").to(device)
processor = CLIPProcessor.from_pretrained("openai/clip-vit-base-patch32")

def get_image_embedding(image_input):
    if isinstance(image_input, str) and image_input.startswith('data:image'):
        image_data = image_input.split(',')[1]
        image_bytes = base64.b64decode(image_data)
        image = Image.open(BytesIO(image_bytes)).convert('RGB')
    elif isinstance(image_input, Image.Image):
        image = image_input
    else:
        image = Image.open(image_input).convert('RGB')
    
    inputs = processor(images=image, return_tensors="pt", padding=True)
    image_features = model.get_image_features(**{k: v.to(device) for k, v in inputs.items()})
    return image_features.cpu().detach().numpy()

def compute_similarity(query_embedding, database_embeddings):
    query_embedding = query_embedding / np.linalg.norm(query_embedding)
    database_embeddings = database_embeddings / np.linalg.norm(database_embeddings, axis=1, keepdims=True)
    return np.dot(database_embeddings, query_embedding.T).flatten()

@app.route('/search', methods=['POST'])
def search():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    try:
        # Process query image
        query_image = Image.open(file).convert('RGB')
        query_embedding = get_image_embedding(query_image)

        # Fetch and process database items
        items_ref = db.collection('items')
        items = items_ref.stream()
        
        database_items = []
        database_embeddings = []
        
        for item in items:
            item_data = item.to_dict()
            if 'imageData' in item_data and item_data['imageData']:
                try:
                    image_data = item_data['imageData'][0]['dataUrl']
                    embedding = get_image_embedding(image_data)
                    database_embeddings.append(embedding.flatten())
                    database_items.append({
                        'id': item.id,
                        'name': item_data.get('name', 'Unnamed Item'),
                        'category': item_data.get('category', 'Uncategorized'),
                        'location': item_data.get('location', 'Unknown location'),
                        'date': item_data.get('date', ''),
                        'description': item_data.get('description', ''),
                        'status': item_data.get('status', 'Available'),
                        'image': image_data,
                        'submitter': item_data.get('submitter', None)
                    })
                except Exception as e:
                    continue

        if not database_items:
            return jsonify({'error': 'No items with images found in database'}), 404

        # Compute and sort similarities
        database_embeddings = np.stack(database_embeddings)
        similarities = compute_similarity(query_embedding.flatten(), database_embeddings)
        results = sorted(zip(similarities, database_items), reverse=True)
        
        return jsonify({
            'results': [
                {
                    'item': item,
                    'similarity': float(score)
                }
                for score, item in results
            ]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def fetch_items_for_export(start_date_dt=None, end_date_dt=None):
    """
    Fetches item data from Firestore for Excel export, filtered by date in Python.
    Handles various date formats stored in Firestore.
    """
    items_ref = db.collection('items')
    # Fetch all documents; filtering will be done in Python
    docs = items_ref.stream()
    data = []

    # Prepare date objects for filtering, using only the date part
    filter_start_date = start_date_dt.date() if start_date_dt else None
    filter_end_date = end_date_dt.date() if end_date_dt else None

    for doc in docs:
        item_data = doc.to_dict()
        
        doc_date_obj = None  # This will be a datetime.date object if parsing succeeds
        doc_date_raw = item_data.get('date')
        date_reported_str = 'N/A'  # For display in Excel

        if doc_date_raw:
            actual_datetime_obj = None
            if isinstance(doc_date_raw, datetime):  # Already a datetime (e.g., from Firestore Timestamp)
                actual_datetime_obj = doc_date_raw
            elif isinstance(doc_date_raw, str):
                # Try parsing various common date and datetime formats
                for fmt in (
                    '%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%m-%d-%Y', '%Y/%m/%d', # Common date formats
                    '%m/%d/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', # Common datetime formats
                    '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%dT%H:%M:%SZ', # ISO formats
                    '%a, %d %b %Y %H:%M:%S %Z' # Example: 'Mon, 12 Jul 2021 14:30:00 GMT' (RFC 822/1123) - less likely from UI
                ):
                    try:
                        actual_datetime_obj = datetime.strptime(doc_date_raw, fmt)
                        break  # Stop on first successful parse
                    except ValueError:
                        continue
            
            if actual_datetime_obj:
                doc_date_obj = actual_datetime_obj.date()  # Use date part for filtering
                date_reported_str = actual_datetime_obj.strftime('%Y-%m-%d %H:%M:%S')  # Use full datetime for display
            elif isinstance(doc_date_raw, str): # If parsing failed, keep original string for display
                date_reported_str = doc_date_raw
            else: # If some other type (e.g. number if mistake was made), convert to string for display
                date_reported_str = str(doc_date_raw)

        # Python-based date filtering
        if filter_start_date and (not doc_date_obj or doc_date_obj < filter_start_date):
            continue  # Skip item if its date is before start_date
        if filter_end_date and (not doc_date_obj or doc_date_obj > filter_end_date):
            continue  # Skip item if its date is after end_date

        # Handle potentially nested submitter information
        submitter_map = item_data.get('submitter') # Get the 'submitter' object/map
        
        submitter_user_id = None
        full_name = 'N/A'
        student_id = 'N/A'

        if isinstance(submitter_map, dict):
            # Try to get userId and denormalized details directly from the submitter_map
            submitter_user_id = submitter_map.get('userId') # Assumes userId is within the submitter map
            full_name = submitter_map.get('displayName', submitter_map.get('full_name', 'N/A'))
            student_id = submitter_map.get('studentId', submitter_map.get('student_id', 'N/A'))
        else:
            # Fallback if submitter is not a map, but a direct ID (less likely given your description)
            submitter_user_id = item_data.get('userId', item_data.get('submitter')) 

        # If full_name or student_id were not found in submitter_map OR if you always want to fetch fresh user data:
        # Proceed to fetch from 'users' collection if a valid userId was found
        if submitter_user_id and isinstance(submitter_user_id, str) and submitter_user_id != 'N/A':
            # Only fetch from users collection if details weren't fully populated from submitter_map
            # or if you have a policy to always fetch the latest user details.
            # For this example, let's assume if full_name is still N/A, we try to fetch.
            if full_name == 'N/A' or student_id == 'N/A': 
                try:
                    user_doc_ref = db.collection('users').document(submitter_user_id)
                    user_doc = user_doc_ref.get()
                    if user_doc.exists:
                        user_details = user_doc.to_dict()
                        # Override with fetched data if it was N/A from submitter_map
                        if full_name == 'N/A':
                            full_name = user_details.get('displayName', user_details.get('name', 'N/A'))
                        if student_id == 'N/A':
                            student_id = user_details.get('studentId', user_details.get('studentID', 'N/A'))
                    else:
                        print(f"User document not found for ID: {submitter_user_id} (item {doc.id})")
                except Exception as e:
                    print(f"Error fetching user details for {submitter_user_id} (item {doc.id}): {e}")
        elif submitter_user_id: # Catches cases where submitter_user_id might be non-string after initial direct get if not map
             print(f"Invalid or non-string submitter_user_id found: {submitter_user_id} for item {doc.id}")

        data.append({
            'Item_ID': doc.id,
            'Item_Name': item_data.get('name', 'N/A'),
            'Category': item_data.get('category', 'N/A'),
            'Location_Found': item_data.get('location', 'N/A'),
            'Submitted_At': date_reported_str, 
            'Status': item_data.get('status', 'N/A'),
            'Description': item_data.get('description', 'N/A'),
            'Student_ID': student_id, 
            'Full_Name': full_name,    
            'Admin_Approved': item_data.get('adminApproval', False)
        })
    
    if not data:
        return pd.DataFrame(columns=['Item_ID', 'Item_Name', 'Category', 'Location_Found', 'Submitted_At', 'Status', 'Description', 'Submitter', 'Full_Name', 'Admin_Approved'])
    return pd.DataFrame(data)

def fetch_users_for_export(start_date_dt=None, end_date_dt=None):
    """
    Fetches user data from Firestore for Excel export.
    Date filtering is currently disabled as 'students' documents lack a date field.
    """
    # Corrected collection name
    users_ref = db.collection('students') 
    query = users_ref
    
    # Date filtering is removed for now as 'students' documents do not have a creation/join date field.
    # If a 'createdAt' or similar timestamp is added to student documents in the future,
    # the following date querying logic can be reinstated:
    # if start_date_dt:
    #     query = query.where('createdAt', '>=', start_date_dt) 
    # if end_date_dt:
    #     query = query.where('createdAt', '<=', end_date_dt)
        
    docs = query.stream()
    data = []
    for doc in docs:
        user = doc.to_dict()
        user_data_for_df = {
            'User_ID': doc.id,
            'Name': user.get('full_name', user.get('name', 'N/A')), 
            'Email': user.get('email', 'N/A'),
            'Student_ID': user.get('student_id', 'N/A'),
            'Year_Level': user.get('year_level', 'N/A'),
            'Status': user.get('status', 'active')
        }
        data.append(user_data_for_df)
        
    # Define column order for the DataFrame
    df_columns = ['User_ID', 'Name', 'Email', 'Student_ID', 'Year_Level', 'Status']
    
    if not data:
        return pd.DataFrame(columns=df_columns)
    return pd.DataFrame(data, columns=df_columns)

def generate_excel_report_to_buffer(data_type, start_date_str=None, end_date_str=None):
    start_date_dt = None
    end_date_dt = None
    if start_date_str:
        try:
            start_date_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
        except ValueError:
            # Allow 'None' or empty string to pass through as no date filter
            if start_date_str.lower() != 'none' and start_date_str != '':
                raise ValueError(f"Invalid start_date format: {start_date_str}. Expected YYYY-MM-DD.")
    if end_date_str:
        try:
            end_date_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            if end_date_str.lower() != 'none' and end_date_str != '':
                raise ValueError(f"Invalid end_date format: {end_date_str}. Expected YYYY-MM-DD.")

    if data_type == 'items':
        df = fetch_items_for_export(start_date_dt, end_date_dt)
        main_sheet_name = 'Items Report'
        title = 'GC Finder: Items Report'
        summary_field = 'Status' 
        summary_title = 'Item Status Summary'
    elif data_type == 'users':
        df = fetch_users_for_export(start_date_dt, end_date_dt) # Dates are now effectively ignored by fetch_users_for_export
        main_sheet_name = 'Users Report'
        title = 'GC Finder: Users Report'
        summary_field = 'Status' 
        summary_title = 'User Status Summary'
    else:
        return None # Should not happen with frontend validation

    if df.empty:
        # Create an empty buffer with a message or just return an empty file
        output_buffer = BytesIO()
        # Optionally, write a message to the Excel file that no data was found
        empty_df_message = pd.DataFrame([{'message': f'No {data_type} found for the selected criteria.'}])
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            empty_df_message.to_excel(writer, sheet_name='No Data', index=False)
        output_buffer.seek(0)
        return output_buffer
        
    output_buffer = BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=main_sheet_name, index=False, startrow=1)
        
        workbook = writer.book
        worksheet = writer.sheets[main_sheet_name]
        
        worksheet['A1'] = title
        last_col_letter = get_column_letter(len(df.columns) if not df.empty else 1)
        worksheet.merge_cells(f'A1:{last_col_letter}1')
        title_cell = worksheet['A1']
        title_cell.font = Font(size=14, bold=True, name='Calibri')
        title_cell.alignment = Alignment(horizontal='center')
        
        for col_idx, column_title in enumerate(df.columns):
            cell = worksheet.cell(row=2, column=col_idx + 1)
            cell.value = column_title
            cell.font = Font(bold=True, name='Calibri')
            cell.alignment = Alignment(horizontal='center')

        status_col_name = summary_field
        if not df.empty and status_col_name in df.columns and not df[status_col_name].dropna().empty:
            status_summary_df = df[status_col_name].value_counts().reset_index()
            status_summary_df.columns = [status_col_name, 'Count']
            
            summary_sheet_name = summary_title
            status_summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False, startrow=1)
            
            summary_ws = writer.sheets[summary_sheet_name]
            summary_ws['A1'] = f"{main_sheet_name} - {summary_title}"
            summary_ws.merge_cells('A1:B1')
            summary_ws['A1'].font = Font(size=14, bold=True, name='Calibri')
            summary_ws['A1'].alignment = Alignment(horizontal='center')

            summary_ws.cell(row=2, column=1).value = status_col_name
            summary_ws.cell(row=2, column=1).font = Font(bold=True, name='Calibri')
            summary_ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
            summary_ws.cell(row=2, column=2).value = 'Count'
            summary_ws.cell(row=2, column=2).font = Font(bold=True, name='Calibri')
            summary_ws.cell(row=2, column=2).alignment = Alignment(horizontal='center')

        for sheetname_iter in writer.sheets:
            current_sheet = writer.sheets[sheetname_iter]
            for col_idx_iter, column_obj in enumerate(current_sheet.columns):
                max_cell_length = 0
                column_letter_val = get_column_letter(col_idx_iter + 1)
                
                header_cell_val = current_sheet.cell(row=1, column=col_idx_iter + 1) # Title row
                if header_cell_val.value and len(str(header_cell_val.value)) > max_cell_length:
                     max_cell_length = len(str(header_cell_val.value))
                
                header_cell_val_2 = current_sheet.cell(row=2, column=col_idx_iter + 1) # Header row
                if header_cell_val_2.value and len(str(header_cell_val_2.value)) > max_cell_length:
                     max_cell_length = len(str(header_cell_val_2.value))

                for cell_obj in column_obj:
                    if cell_obj.value:
                        cell_len_val = len(str(cell_obj.value))
                        if cell_len_val > max_cell_length:
                            max_cell_length = cell_len_val
                current_sheet.column_dimensions[column_letter_val].width = max_cell_length + 3 if max_cell_length > 0 else 12
    
    output_buffer.seek(0)
    return output_buffer

@app.route('/api/export', methods=['GET'])
def export_data_route():
    data_type = request.args.get('type')
    start_date_str = request.args.get('startDate')
    end_date_str = request.args.get('endDate')

    if not data_type:
        return jsonify({"error": "Missing 'type' parameter (should be 'items' or 'users')"}), 400
    if data_type not in ['items', 'users']:
        return jsonify({"error": "Invalid 'type' parameter (should be 'items' or 'users')"}), 400

    try:
        excel_buffer = generate_excel_report_to_buffer(
            data_type=data_type, 
            start_date_str=start_date_str, 
            end_date_str=end_date_str
        )
        
        filename_prefix = data_type
        date_range_str = "all_time"
        if start_date_str and end_date_str:
            date_range_str = f"{start_date_str}_to_{end_date_str}"
        elif start_date_str:
            date_range_str = f"from_{start_date_str}"
        elif end_date_str:
            date_range_str = f"up_to_{end_date_str}"
        
        output_filename = f"GCFinder_{filename_prefix}_report_{date_range_str}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_filename 
        )
    except ValueError as ve:
        app.logger.error(f"ValueError during report generation: {ve}")
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        app.logger.error(f"Unexpected error during report generation: {e}", exc_info=True)
        return jsonify({"error": "An unexpected error occurred. Please check server logs."}), 500

if __name__ == '__main__':
    app.run(debug=True)
